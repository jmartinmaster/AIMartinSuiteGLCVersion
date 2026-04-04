# The Martin Suite (GLC Edition)
# Copyright (C) 2026 Jamie Martin
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.

import openpyxl
import json
import os
import re
import shutil
import sys
from datetime import date, datetime

from modules.downtime_codes import get_code_number, normalize_code_value

__module_name__ = "Data Handler"
__version__ = "1.1.0"

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

def external_path(relative_path):
    """Get path to external file (Write-Enabled)."""
    return os.path.join(os.path.abspath("."), relative_path)

class DataHandler:
    def __init__(self, config_name="layout_config.json"):
        # Prefer the local config so Layout Manager changes are reflected in import/export.
        self.config_path = external_path(config_name)
        if not os.path.exists(self.config_path):
            self.config_path = resource_path(config_name)
        
        with open(self.config_path, 'r', encoding='utf-8') as f:
            self.config = json.load(f)

        self.settings = self.load_settings()
        
        # User data folders stay in the local directory (not internal to the exe)
        self.pending_dir = "data/pending"
        os.makedirs(self.pending_dir, exist_ok=True)

    def load_settings(self):
        settings_path = external_path("settings.json")
        settings = {
            "export_directory": "exports",
            "organize_exports_by_date": True,
            "default_export_prefix": "Disamatic Production Sheet",
        }
        if os.path.exists(settings_path):
            try:
                with open(settings_path, 'r', encoding='utf-8') as handle:
                    loaded = json.load(handle)
                if isinstance(loaded, dict):
                    settings.update(loaded)
            except Exception:
                pass
        return settings

    def parse_export_date(self, raw_date):
        date_text = str(raw_date or "").strip()
        if not date_text:
            return None

        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%m%d%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(date_text, fmt)
            except ValueError:
                continue
        return None

    def format_cell_value(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%Y")
        if isinstance(value, date):
            return value.strftime("%m/%d/%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        return value

    def format_header_value(self, field_id, value, number_format=None):
        formatted = self.format_cell_value(value)
        if isinstance(value, (int, float)) and number_format and '%' in str(number_format):
            return f"{value * 100:.0f}%"
        if field_id == "cast_date":
            text = str(formatted or "").strip()
            if not text:
                return ""
            digits = "".join(ch for ch in text if ch.isdigit())
            if digits:
                return digits.zfill(3)[-3:]
        return formatted

    def compute_cast_date(self, raw_date):
        parsed = self.parse_export_date(raw_date)
        if parsed is None:
            return ""
        return f"{parsed.timetuple().tm_yday:03}"

    def evaluate_formula_cell(self, workbook, worksheet, formula, cache=None):
        if not isinstance(formula, str) or not formula.startswith("="):
            return formula

        cache = cache if cache is not None else {}
        cache_key = (worksheet.title, formula)
        if cache_key in cache:
            return cache[cache_key]

        expression = formula[1:].strip()

        def resolve_reference(token):
            if '!' in token:
                sheet_name, cell_ref = token.split('!', 1)
                sheet_name = sheet_name.strip("'")
                target_ws = workbook[sheet_name]
            else:
                cell_ref = token
                target_ws = worksheet
            return self.resolve_import_cell_value(workbook, target_ws, cell_ref, cache)

        def replace_sum(match):
            args = [part.strip() for part in match.group(1).split(',') if part.strip()]
            total = 0
            for arg in args:
                if ':' in arg:
                    if '!' in arg:
                        sheet_name, cell_range = arg.split('!', 1)
                        target_ws = workbook[sheet_name.strip("'")]
                    else:
                        cell_range = arg
                        target_ws = worksheet
                    for row in target_ws[cell_range]:
                        for cell in row:
                            value = self.resolve_import_cell_value(workbook, target_ws, cell.coordinate, cache)
                            if isinstance(value, (int, float)):
                                total += value
                            elif value not in (None, ""):
                                try:
                                    total += float(value)
                                except Exception:
                                    pass
                else:
                    value = resolve_reference(arg)
                    if isinstance(value, (int, float)):
                        total += value
                    elif value not in (None, ""):
                        try:
                            total += float(value)
                        except Exception:
                            pass
            return str(total)

        expression = re.sub(r'SUM\(([^\)]*)\)', replace_sum, expression, flags=re.IGNORECASE)

        ref_pattern = re.compile(r"(?<![A-Z0-9_'])((?:'[^']+'!)?[A-Z]+\d+)")

        def replace_ref(match):
            token = match.group(1)
            value = resolve_reference(token)
            if value in (None, ""):
                return "0"
            if isinstance(value, str):
                try:
                    return str(float(value))
                except Exception:
                    return "0"
            return str(value)

        expression = ref_pattern.sub(replace_ref, expression)
        try:
            result = eval(expression, {"__builtins__": None}, {})
        except Exception:
            result = None
        cache[cache_key] = result
        return result

    def resolve_import_cell_value(self, workbook, worksheet, cell_ref, cache=None):
        cache = cache if cache is not None else {}
        cell = worksheet[cell_ref]
        value = cell.value
        if isinstance(value, str) and value.startswith('='):
            return self.evaluate_formula_cell(workbook, worksheet, value, cache)
        return value

    def normalize_time_value(self, value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        digits = "".join(ch for ch in text if ch.isdigit())
        if not digits:
            return ""
        return digits.zfill(4)[-4:]

    def parse_total_minutes(self, value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return max(0, int(value))
        text = str(value).strip()
        if not text:
            return None
        digits = "".join(ch for ch in text if ch.isdigit())
        if not digits:
            return None
        return max(0, int(digits))

    def calculate_duration_minutes(self, start_value, stop_value):
        start_text = self.normalize_time_value(start_value)
        stop_text = self.normalize_time_value(stop_value)
        if len(start_text) != 4 or len(stop_text) != 4:
            return None

        start_minutes = int(start_text[:2]) * 60 + int(start_text[2:])
        stop_minutes = int(stop_text[:2]) * 60 + int(stop_text[2:])
        if stop_minutes < start_minutes:
            stop_minutes += 24 * 60
        return stop_minutes - start_minutes

    def calculate_stop_time(self, start_value, total_minutes_value):
        start_text = self.normalize_time_value(start_value)
        total_minutes = self.parse_total_minutes(total_minutes_value)
        if len(start_text) != 4 or total_minutes is None:
            return ""

        start_minutes = int(start_text[:2]) * 60 + int(start_text[2:])
        stop_minutes = (start_minutes + total_minutes) % (24 * 60)
        return f"{stop_minutes // 60:02}{stop_minutes % 60:02}"

    def get_export_directory(self, raw_date):
        base_dir = str(self.settings.get("export_directory", "exports") or "exports").strip() or "exports"
        target_dir = external_path(base_dir)
        if self.settings.get("organize_exports_by_date", True):
            export_date = self.parse_export_date(raw_date)
            if export_date is not None:
                target_dir = os.path.join(target_dir, export_date.strftime("%Y"), export_date.strftime("%m"))
        os.makedirs(target_dir, exist_ok=True)
        return target_dir

    def calculate_formula(self, formula_str, data_context):
    # """
        #Takes a string like '{molds} * 0.95' and replaces {molds} 
        #With the actual value from the UI entries.
        #"""
        try:
            # Replace placeholders {id} with actual values from data_context
            for key, value in data_context.items():
                placeholder = "{" + key + "}"
                if placeholder in formula_str:
                    # Default to 0 if the field is empty
                    val = value if value and str(value).strip() else "0"
                    formula_str = formula_str.replace(placeholder, str(val))
            
            # Clean the string for safety and evaluate
            # We only allow numbers and basic math operators
            return eval(formula_str, {"__builtins__": None}, {})
        except Exception as e:
            print(f"Math Error in formula '{formula_str}': {e}")
            return 0

    def get_production_column(self, column_map, key, fallback=None):
        value = column_map.get(key, fallback)
        text = str(value or "").strip()
        return text or None

    def normalize_header_label(self, value):
        return " ".join(str(value or "").strip().lower().split())

    def detect_production_columns(self, worksheet, configured_columns, start_row):
        resolved_columns = dict(configured_columns)
        header_row = max(1, int(start_row) - 1)
        label_map = {
            "shop order": "shop_order",
            "part number": "part_number",
            "molds": "molds",
        }

        for column_index in range(1, worksheet.max_column + 1):
            raw_value = worksheet.cell(row=header_row, column=column_index).value
            normalized = self.normalize_header_label(raw_value)
            field_name = label_map.get(normalized)
            if field_name:
                resolved_columns[field_name] = openpyxl.utils.get_column_letter(column_index)

        return resolved_columns

    def export_to_template(self, ui_data, shift, date_str):
        clean_date = date_str.replace("/", "")
        export_prefix = str(self.settings.get("default_export_prefix", "Disamatic Production Sheet") or "Disamatic Production Sheet").strip()
        filename = f"{export_prefix} {shift}{clean_date}.xlsx"
        target_path = os.path.join(self.get_export_directory(date_str), filename)

        # CRITICAL: Use resource_path for the template path defined in your JSON
        # If your JSON says "templates/production.xlsx", this finds it in the build
        full_template_path = resource_path(self.config['template_path'])
        
        shutil.copy(full_template_path, target_path)
        wb = openpyxl.load_workbook(target_path)
        ws = wb.active

        # 1. Map Header
        for field in self.config['header_fields']:
            cell_coord = field.get('cell')
            val = ui_data['header'].get(field['id'])
            if cell_coord and field.get('export_enabled', True):
                cell = ws[cell_coord]
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell, MergedCell):
                    for range_ in ws.merged_cells.ranges:
                        if cell_coord in range_:
                            ws.cell(range_.min_row, range_.min_col).value = val
                            break
                else:
                    cell.value = val

        # 2. Map Production Rows
        p_map = self.config['production_mapping']
        p_start = p_map['start_row']
        p_cols = p_map['columns']
        for i, row_data in enumerate(ui_data['production']):
            curr_row = p_start + i
            ws[f"{p_cols['shop_order']}{curr_row}"] = row_data.get('shop_order')
            ws[f"{p_cols['part_number']}{curr_row}"] = row_data.get('part_number')
            ws[f"{p_cols['molds']}{curr_row}"] = row_data.get('molds')

        # 3. Map Downtime Rows
        d_map = self.config['downtime_mapping']
        d_start = d_map['start_row']
        d_cols = d_map['columns']

        for i, row_data in enumerate(ui_data['downtime']):
            curr_row = d_start + i
            short_code = get_code_number(row_data.get('code', ""))
            duration_minutes = self.calculate_duration_minutes(row_data.get('start'), row_data.get('stop'))
            if duration_minutes is None:
                duration_minutes = self.parse_total_minutes(str(row_data.get('time_calc', '')).replace(' min', ''))
            
            ws[f"{d_cols['start']}{curr_row}"] = row_data.get('start')
            ws[f"{d_cols['stop']}{curr_row}"] = duration_minutes
            ws[f"{d_cols['code']}{curr_row}"] = short_code 
            ws[f"{d_cols['cause']}{curr_row}"] = row_data.get('cause')

        wb.save(target_path)
        return target_path

    # ... [import_from_excel method remains the same] ...

    # Import Production
    def import_from_excel(self, file_path):
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            formula_wb = openpyxl.load_workbook(file_path, data_only=False)
            ws = wb.active
            formula_ws = formula_wb.active
            formula_cache = {}
            
            data = {"header": {}, "production": [], "downtime": []}

            # 2. Import Header (Date, Shift, etc.)
            for field in self.config['header_fields']:
                cell_coord = field.get('cell')
                if cell_coord and field.get('import_enabled', True):
                    value = ws[cell_coord].value
                    if value is None:
                        value = self.resolve_import_cell_value(formula_wb, formula_ws, cell_coord, formula_cache)
                    number_format = formula_ws[cell_coord].number_format if cell_coord else None
                    data["header"][field['id']] = self.format_header_value(field['id'], value, number_format)

            if not data["header"].get("cast_date"):
                data["header"]["cast_date"] = self.compute_cast_date(data["header"].get("date"))

            # 3. Import Production Lines (Shop Order, Part #, Molds)
            p_map = self.config['production_mapping']
            p_cols = self.detect_production_columns(formula_ws, p_map['columns'], p_map['start_row'])
            for i in range(50): # Check up to 50 rows
                row_idx = p_map['start_row'] + i
                shop_order = ws[f"{p_cols['shop_order']}{row_idx}"].value
                if shop_order is None:
                    shop_order = self.resolve_import_cell_value(formula_wb, formula_ws, f"{p_cols['shop_order']}{row_idx}", formula_cache)
                
                # If the shop order cell is empty, we've reached the end of the list
                if not shop_order:
                    break

                part_number = ws[f"{p_cols['part_number']}{row_idx}"].value
                if part_number is None:
                    part_number = self.resolve_import_cell_value(formula_wb, formula_ws, f"{p_cols['part_number']}{row_idx}", formula_cache)

                molds = ws[f"{p_cols['molds']}{row_idx}"].value
                if molds is None:
                    molds = self.resolve_import_cell_value(formula_wb, formula_ws, f"{p_cols['molds']}{row_idx}", formula_cache)

                data["production"].append({
                    "shop_order": self.format_cell_value(shop_order),
                    "part_number": self.format_cell_value(part_number),
                    "molds": self.format_cell_value(molds)
                })

            # 4. Import Downtime (Start, Total Minutes, Code, Cause)
            d_map = self.config['downtime_mapping']
            d_cols = d_map['columns']
            for i in range(25): # Check up to 25 downtime rows
                row_idx = d_map['start_row'] + i
                start_time = ws[f"{d_cols['start']}{row_idx}"].value
                if start_time is None:
                    start_time = self.resolve_import_cell_value(formula_wb, formula_ws, f"{d_cols['start']}{row_idx}", formula_cache)
                
                if not start_time:
                    break

                # Code conversion: UI needs "1 Misc..." but Excel only has "1"
                raw_val = ws[f"{d_cols['code']}{row_idx}"].value
                if raw_val is None:
                    raw_val = self.resolve_import_cell_value(formula_wb, formula_ws, f"{d_cols['code']}{row_idx}", formula_cache)
                full_code = normalize_code_value(raw_val)
                total_minutes = ws[f"{d_cols['stop']}{row_idx}"].value
                if total_minutes is None:
                    total_minutes = self.resolve_import_cell_value(formula_wb, formula_ws, f"{d_cols['stop']}{row_idx}", formula_cache)
                cause = ws[f"{d_cols['cause']}{row_idx}"].value
                if cause is None:
                    cause = self.resolve_import_cell_value(formula_wb, formula_ws, f"{d_cols['cause']}{row_idx}", formula_cache)

                data["downtime"].append({
                    "start": self.format_cell_value(start_time),
                    "stop": self.calculate_stop_time(start_time, total_minutes),
                    "code": full_code,
                    "cause": self.format_cell_value(cause)
                })

            return data