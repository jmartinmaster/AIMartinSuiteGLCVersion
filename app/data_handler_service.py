# Production Logging Center (GLC Edition)
# Copyright (C) 2026 Jamie Martin
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
import json
import math
import os
import re
import shutil
from datetime import date, datetime

import openpyxl

from app.downtime_codes import get_code_number, normalize_code_value
from app.form_definition_registry import FormDefinitionRegistry
from app.production_log_roles import HEADER_DERIVED_ROLES, PRODUCTION_IMPORT_LABEL_ROLES, get_default_row_field_id, normalize_role_name, resolve_header_field_role, resolve_row_field_role
from app.safe_expression import SafeExpressionEvaluator
from app.utils import ensure_external_directory, external_path, resource_path

__module_name__ = "Data Handler"
__version__ = "1.1.6"

DEFAULT_SHIFT_TIME_SETTINGS = {
    "shift_total_rounding": "nearest",
    "shift_1_anchor_mode": "start",
    "shift_1_reference_time": "0600",
    "shift_2_anchor_mode": "midpoint",
    "shift_2_reference_time": "1800",
    "shift_3_anchor_mode": "end",
    "shift_3_reference_time": "0600",
}
ROW_SECTION_CONFIG = {
    "production": {
        "field_key": "production_row_fields",
        "mapping_key": "production_mapping",
        "default_max_rows": 50,
    },
    "downtime": {
        "field_key": "downtime_row_fields",
        "mapping_key": "downtime_mapping",
        "default_max_rows": 25,
    },
}
SAFE_EXPRESSION_EVALUATOR = SafeExpressionEvaluator()


class DataHandlerService:
    def __init__(self, form_id=None):
        self.form_registry = FormDefinitionRegistry()
        self.form_info = self.form_registry.get_form(form_id)
        self.config_path = self.form_info["load_path"]
        with open(self.config_path, "r", encoding="utf-8") as handle:
            self.config = self._normalize_layout_config(json.load(handle))
        self.settings = self.load_settings()
        self.pending_dir = ensure_external_directory("data/pending")

    def _normalize_layout_config(self, config):
        normalized = dict(config) if isinstance(config, dict) else {}
        normalized["header_fields"] = self._normalize_header_field_configs(normalized.get("header_fields"))
        for section_name, section_config in ROW_SECTION_CONFIG.items():
            field_key = section_config.get("field_key")
            normalized[field_key] = self._normalize_row_field_configs(normalized.get(field_key), section_name)
        return normalized

    def _normalize_header_field_configs(self, configured_fields):
        if not isinstance(configured_fields, list):
            return []

        normalized_fields = []
        seen_ids = set()
        for field in configured_fields:
            if not isinstance(field, dict):
                continue
            field_id = str(field.get("id", "")).strip()
            if not field_id or field_id in seen_ids:
                continue
            normalized_field = dict(field)
            role_name = resolve_header_field_role(field_id, normalized_field.get("role"))
            if role_name:
                normalized_field["role"] = role_name
            else:
                normalized_field.pop("role", None)
            normalized_fields.append(normalized_field)
            seen_ids.add(field_id)
        return normalized_fields

    def _normalize_row_field_configs(self, configured_fields, section_name):
        if not isinstance(configured_fields, list):
            return []

        normalized_fields = []
        seen_ids = set()
        for field in configured_fields:
            if not isinstance(field, dict):
                continue
            field_id = str(field.get("id", "")).strip()
            if not field_id or field_id in seen_ids:
                continue
            normalized_field = dict(field)
            role_name = resolve_row_field_role(section_name, field_id, normalized_field.get("role"))
            if role_name:
                normalized_field["role"] = role_name
            else:
                normalized_field.pop("role", None)
            normalized_fields.append(normalized_field)
            seen_ids.add(field_id)
        return normalized_fields

    def load_settings(self):
        settings_path = external_path("settings.json")
        settings = {
            "export_directory": "exports",
            "organize_exports_by_date": True,
            "default_export_prefix": "Disamatic Production Sheet",
        }
        if os.path.exists(settings_path):
            try:
                with open(settings_path, "r", encoding="utf-8") as handle:
                    loaded = json.load(handle)
                if isinstance(loaded, dict):
                    settings.update(loaded)
            except Exception:
                pass
        return settings

    def parse_export_date(self, raw_date):
        date_text = str(raw_date or "").strip()
        if not date_text:
            return None
        for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%m%d%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(date_text, fmt)
            except ValueError:
                continue
        return None

    def format_cell_value(self, value):
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%m/%d/%Y")
        if isinstance(value, date):
            return value.strftime("%m/%d/%Y")
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int):
            return str(value)
        return value

    def format_header_value(self, field_id, value, number_format=None):
        formatted = self.format_cell_value(value)
        if isinstance(value, (int, float)) and number_format and "%" in str(number_format):
            return f"{value * 100:.0f}%"
        if self.get_header_field_role(field_id) == "cast_date":
            text = str(formatted or "").strip()
            if not text:
                return ""
            digits = "".join(ch for ch in text if ch.isdigit())
            if digits:
                return digits.zfill(3)[-3:]
        return formatted

    def get_header_fields(self):
        return self.config.get("header_fields", [])

    def get_header_field_config(self, field_id):
        for field in self.get_header_fields():
            if field.get("id") == field_id:
                return field
        return {}

    def get_header_field_role(self, field_id):
        field_config = self.get_header_field_config(field_id)
        return resolve_header_field_role(field_id, field_config.get("role"))

    def get_header_field_id_by_role(self, role_name, fallback_id=None):
        normalized_role = normalize_role_name(role_name)
        for field in self.get_header_fields():
            if resolve_header_field_role(field.get("id"), field.get("role")) == normalized_role:
                return field.get("id")
        return fallback_id

    def get_header_value_by_role(self, header_data, role_name, fallback_id=None, default=""):
        if not isinstance(header_data, dict):
            return default
        field_id = self.get_header_field_id_by_role(role_name, fallback_id=fallback_id)
        if field_id and field_id in header_data:
            return header_data.get(field_id, default)
        if fallback_id and fallback_id in header_data:
            return header_data.get(fallback_id, default)
        return default

    def get_header_field_ids(self):
        return [field.get("id") for field in self.get_header_fields() if field.get("id")]

    def get_section_field_configs(self, section_name):
        section_config = ROW_SECTION_CONFIG.get(section_name, {})
        field_key = section_config.get("field_key")
        if not field_key:
            return []
        fields = self.config.get(field_key, [])
        return [dict(field) for field in fields if isinstance(field, dict) and field.get("id")]

    def get_section_field_role(self, section_name, field_id):
        for field in self.get_section_field_configs(section_name):
            if field.get("id") == field_id:
                return resolve_row_field_role(section_name, field_id, field.get("role"))
        return resolve_row_field_role(section_name, field_id)

    def get_section_field_id_by_role(self, section_name, role_name, fallback_id=None):
        normalized_role = normalize_role_name(role_name)
        for field in self.get_section_field_configs(section_name):
            if resolve_row_field_role(section_name, field.get("id"), field.get("role")) == normalized_role:
                return field.get("id")
        return fallback_id

    def get_row_value_by_role(self, section_name, row_data, role_name, fallback_id=None, default=None):
        if not isinstance(row_data, dict):
            return default
        field_id = self.get_section_field_id_by_role(section_name, role_name, fallback_id=fallback_id)
        if field_id and field_id in row_data:
            return row_data.get(field_id, default)
        if fallback_id and fallback_id in row_data:
            return row_data.get(fallback_id, default)
        return default

    def get_section_field_config(self, section_name, field_id):
        for field in self.get_section_field_configs(section_name):
            if field.get("id") == field_id:
                return field
        return {}

    def get_open_row_field_ids(self, section_name):
        return [field.get("id") for field in self.get_section_field_configs(section_name) if field.get("open_row_trigger")]

    def format_numeric_text(self, value, allow_decimal=False):
        text = str(value or "").strip()
        if not text:
            return ""
        try:
            numeric_value = float(text)
        except Exception:
            return text
        if numeric_value.is_integer():
            return str(int(numeric_value))
        if allow_decimal:
            return f"{numeric_value:.2f}".rstrip("0").rstrip(".")
        return str(int(round(numeric_value)))

    def normalize_date_text(self, value):
        text = str(value or "").strip()
        if not text:
            return ""
        parsed = self.parse_export_date(text)
        if parsed is None:
            return text
        return parsed.strftime("%m/%d/%Y")

    def _normalize_calculation_settings(self, calculation_settings=None):
        settings = dict(DEFAULT_SHIFT_TIME_SETTINGS)
        if isinstance(calculation_settings, dict):
            settings.update(calculation_settings)
        return settings

    def _round_minutes_value(self, value, rounding_mode):
        try:
            normalized_value = float(value)
        except Exception:
            return 0
        if normalized_value <= 0:
            return 0
        normalized_mode = str(rounding_mode or "nearest").strip().lower()
        if normalized_mode == "ceil":
            return int(math.ceil(normalized_value))
        if normalized_mode == "floor":
            return int(math.floor(normalized_value))
        return int(round(normalized_value))

    def _parse_compact_time(self, value):
        digits = "".join(ch for ch in str(value or "").strip() if ch.isdigit())
        if not digits:
            return None
        digits = digits.zfill(4)[-4:]
        hours = int(digits[:2])
        minutes = int(digits[2:])
        if hours > 23 or minutes > 59:
            return None
        return (hours * 60) + minutes

    def _format_compact_time(self, total_minutes):
        normalized_minutes = int(total_minutes) % (24 * 60)
        hours, minutes = divmod(normalized_minutes, 60)
        return f"{hours:02d}{minutes:02d}"

    def _normalize_clock_formula_result(self, value):
        if value in (None, ""):
            return ""
        if isinstance(value, (int, float)):
            return self._format_compact_time(int(value))
        parsed = self._parse_compact_time(value)
        if parsed is not None:
            return self._format_compact_time(parsed)
        return str(value).strip()

    def _normalize_shift_key(self, shift_value):
        normalized = str(shift_value or "").strip()
        return normalized if normalized in {"1", "2", "3"} else ""

    def compute_shift_window(self, shift_value, raw_hours, calculation_settings=None):
        settings = self._normalize_calculation_settings(calculation_settings)
        shift_key = self._normalize_shift_key(shift_value)
        if not shift_key:
            return {"start_time": "", "end_time": "", "target_time": ""}

        total_minutes = self.calculate_shift_total_minutes(raw_hours, calculation_settings=calculation_settings)
        if total_minutes <= 0:
            return {"start_time": "", "end_time": "", "target_time": ""}

        anchor_mode = str(settings.get(f"shift_{shift_key}_anchor_mode", "start") or "start").strip().lower()
        anchor_minutes = self._parse_compact_time(settings.get(f"shift_{shift_key}_reference_time", ""))
        if anchor_minutes is None:
            return {"start_time": "", "end_time": "", "target_time": ""}

        if anchor_mode == "end":
            end_minutes = anchor_minutes
            start_minutes = end_minutes - total_minutes
        elif anchor_mode == "midpoint":
            midpoint_offset = total_minutes / 2.0
            start_minutes = int(round(anchor_minutes - midpoint_offset))
            end_minutes = start_minutes + total_minutes
        else:
            start_minutes = anchor_minutes
            end_minutes = start_minutes + total_minutes

        formulas = settings.get("formulas", {}) if isinstance(settings, dict) else {}
        formula_context = {
            "shift_anchor_mode": anchor_mode,
            "anchor_minutes": anchor_minutes,
            "shift_total_minutes": total_minutes,
            "default_start_minutes": start_minutes,
            "default_end_minutes": end_minutes,
            "day_minutes": 24 * 60,
        }
        start_formula = str(formulas.get("shift_start_time") or "").strip()
        end_formula = str(formulas.get("shift_end_time") or "").strip()
        start_time = self._normalize_clock_formula_result(
            self.evaluate_expression_formula(start_formula, formula_context, default=self._format_compact_time(start_minutes))
        )
        end_time = self._normalize_clock_formula_result(
            self.evaluate_expression_formula(end_formula, formula_context, default=self._format_compact_time(end_minutes))
        )

        target_time_field_id = self.get_header_field_id_by_role("target_time", fallback_id="target_time")
        field_cfg = self.get_header_field_config(target_time_field_id)
        suffix = field_cfg.get("suffix", " min")
        return {
            "start_time": start_time,
            "end_time": end_time,
            "target_time": f"{total_minutes}{suffix}",
        }

    def compute_target_time(self, raw_hours, calculation_settings=None):
        total_minutes = self.calculate_shift_total_minutes(raw_hours, calculation_settings=calculation_settings)
        target_time_field_id = self.get_header_field_id_by_role("target_time", fallback_id="target_time")
        field_cfg = self.get_header_field_config(target_time_field_id)
        suffix = field_cfg.get("suffix", " min")
        return f"{total_minutes}{suffix}" if total_minutes > 0 else ""

    def calculate_shift_total_minutes(self, raw_hours, calculation_settings=None):
        settings = self._normalize_calculation_settings(calculation_settings)
        try:
            hours_value = float(raw_hours or 0)
        except Exception:
            return 0

        formulas = settings.get("formulas", {}) if isinstance(settings, dict) else {}
        formula_text = str(formulas.get("shift_total_minutes") or "").strip()
        if formula_text:
            result = self.evaluate_expression_formula(
                formula_text,
                {
                    "hours": hours_value,
                    "shift_total_rounding": settings.get("shift_total_rounding", "nearest"),
                },
                default=0,
            )
            try:
                return max(0, int(float(result or 0)))
            except Exception:
                return 0
        return self._round_minutes_value(hours_value * 60, settings.get("shift_total_rounding", "nearest"))

    def normalize_target_time_text(self, value):
        total_minutes = self.parse_total_minutes(value)
        if total_minutes is None or total_minutes <= 0:
            return ""
        target_time_field_id = self.get_header_field_id_by_role("target_time", fallback_id="target_time")
        field_cfg = self.get_header_field_config(target_time_field_id)
        suffix = field_cfg.get("suffix", " min")
        return f"{total_minutes}{suffix}"

    def normalize_header_field_value(self, field_id, value, header_data=None, calculation_settings=None):
        header_data = header_data or {}
        text = str(value or "").strip()
        field_role = self.get_header_field_role(field_id)
        date_field_id = self.get_header_field_id_by_role("log_date", fallback_id="date")
        shift_field_id = self.get_header_field_id_by_role("shift_number", fallback_id="shift")
        hours_field_id = self.get_header_field_id_by_role("shift_hours", fallback_id="hours")
        if field_role == "log_date":
            return self.normalize_date_text(text)
        if field_role == "shift_hours":
            return self.format_numeric_text(text, allow_decimal=True)
        if field_role in {"shift_number", "goal_rate", "ret_south", "ret_north", "total_molds"}:
            return self.format_numeric_text(text, allow_decimal=False)
        if field_role == "cast_date":
            computed = self.compute_cast_date(header_data.get(date_field_id))
            return computed or self.format_header_value(field_id, text)
        if field_role in {"shift_start_time", "shift_end_time"}:
            shift_window = self.compute_shift_window(
                header_data.get(shift_field_id),
                header_data.get(hours_field_id),
                calculation_settings=calculation_settings,
            )
            result_key = "start_time" if field_role == "shift_start_time" else "end_time"
            return shift_window.get(result_key, "")
        if field_role == "target_time":
            computed = self.compute_target_time(header_data.get(hours_field_id), calculation_settings=calculation_settings)
            return computed or self.normalize_target_time_text(text)
        return text

    def normalize_header_data(self, header_data, calculation_settings=None):
        normalized = {}
        raw_header = {field_id: str(header_data.get(field_id, "") or "").strip() for field_id in self.get_header_field_ids()}
        for field_id in self.get_header_field_ids():
            if self.get_header_field_role(field_id) in HEADER_DERIVED_ROLES:
                continue
            normalized[field_id] = self.normalize_header_field_value(
                field_id,
                raw_header.get(field_id, ""),
                {**raw_header, **normalized},
                calculation_settings=calculation_settings,
            )
        cast_date_field_id = self.get_header_field_id_by_role("cast_date", fallback_id="cast_date")
        if cast_date_field_id in raw_header:
            normalized[cast_date_field_id] = self.normalize_header_field_value(
                cast_date_field_id,
                raw_header.get(cast_date_field_id, ""),
                {**raw_header, **normalized},
                calculation_settings=calculation_settings,
            )
        for derived_role in ("shift_start_time", "shift_end_time", "target_time"):
            derived_field_id = self.get_header_field_id_by_role(derived_role)
            if derived_field_id not in raw_header:
                continue
            normalized[derived_field_id] = self.normalize_header_field_value(
                derived_field_id,
                raw_header.get(derived_field_id, ""),
                {**raw_header, **normalized},
                calculation_settings=calculation_settings,
            )
        return normalized

    def compute_cast_date(self, raw_date):
        parsed = self.parse_export_date(raw_date)
        if parsed is None:
            return ""
        return f"{parsed.timetuple().tm_yday:03}"

    def evaluate_formula_cell(self, workbook, worksheet, formula, cache=None):
        if not isinstance(formula, str) or not formula.startswith("="):
            return formula
        cache = cache if cache is not None else {}
        cache_key = (worksheet.title, formula)
        if cache_key in cache:
            return cache[cache_key]
        expression = formula[1:].strip()

        def resolve_reference(token):
            if "!" in token:
                sheet_name, cell_ref = token.split("!", 1)
                sheet_name = sheet_name.strip("'")
                target_ws = workbook[sheet_name]
            else:
                cell_ref = token
                target_ws = worksheet
            return self.resolve_import_cell_value(workbook, target_ws, cell_ref, cache)

        def replace_sum(match):
            args = [part.strip() for part in match.group(1).split(",") if part.strip()]
            total = 0
            for arg in args:
                if ":" in arg:
                    if "!" in arg:
                        sheet_name, cell_range = arg.split("!", 1)
                        target_ws = workbook[sheet_name.strip("'")]
                    else:
                        cell_range = arg
                        target_ws = worksheet
                    for row in target_ws[cell_range]:
                        for cell in row:
                            value = self.resolve_import_cell_value(workbook, target_ws, cell.coordinate, cache)
                            if isinstance(value, (int, float)):
                                total += value
                            elif value not in (None, ""):
                                try:
                                    total += float(value)
                                except Exception:
                                    pass
                else:
                    value = resolve_reference(arg)
                    if isinstance(value, (int, float)):
                        total += value
                    elif value not in (None, ""):
                        try:
                            total += float(value)
                        except Exception:
                            pass
            return str(total)

        expression = re.sub(r"SUM\(([^\)]*)\)", replace_sum, expression, flags=re.IGNORECASE)
        ref_pattern = re.compile(r"(?<![A-Z0-9_'])((?:'[^']+'!)?[A-Z]+\d+)")

        def replace_ref(match):
            token = match.group(1)
            value = resolve_reference(token)
            if value in (None, ""):
                return "0"
            if isinstance(value, str):
                try:
                    return str(float(value))
                except Exception:
                    return "0"
            return str(value)

        expression = ref_pattern.sub(replace_ref, expression)
        result = self._evaluate_safe_expression(expression, default=None)
        cache[cache_key] = result
        return result

    def resolve_import_cell_value(self, workbook, worksheet, cell_ref, cache=None):
        cache = cache if cache is not None else {}
        cell = worksheet[cell_ref]
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            return self.evaluate_formula_cell(workbook, worksheet, value, cache)
        return value

    def normalize_time_value(self, value):
        if value is None:
            return ""
        text = str(value).strip()
        if not text:
            return ""
        digits = "".join(ch for ch in text if ch.isdigit())
        if not digits:
            return ""
        return digits.zfill(4)[-4:]

    def parse_total_minutes(self, value):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return max(0, int(value))
        text = str(value).strip()
        if not text:
            return None
        digits = "".join(ch for ch in text if ch.isdigit())
        if not digits:
            return None
        return max(0, int(digits))

    def calculate_duration_minutes(self, start_value, stop_value, allow_overnight=True):
        start_text = self.normalize_time_value(start_value)
        stop_text = self.normalize_time_value(stop_value)
        if len(start_text) != 4 or len(stop_text) != 4:
            return None
        start_minutes = int(start_text[:2]) * 60 + int(start_text[2:])
        stop_minutes = int(stop_text[:2]) * 60 + int(stop_text[2:])
        if stop_minutes < start_minutes:
            if not allow_overnight:
                return None
            stop_minutes += 24 * 60
        return stop_minutes - start_minutes

    def calculate_runtime_downtime_minutes(self, start_value, stop_value, calculation_settings=None, fallback_label=None):
        settings = self._normalize_calculation_settings(calculation_settings)
        start_text = self.normalize_time_value(start_value)
        stop_text = self.normalize_time_value(stop_value)
        if len(start_text) != 4 or len(stop_text) != 4:
            if fallback_label is not None:
                return self.parse_total_minutes(fallback_label)
            return None

        start_minutes = int(start_text[:2]) * 60 + int(start_text[2:])
        stop_minutes = int(stop_text[:2]) * 60 + int(stop_text[2:])
        formulas = settings.get("formulas", {}) if isinstance(settings, dict) else {}
        formula_text = str(formulas.get("downtime_minutes") or "").strip()
        if formula_text:
            result = self.evaluate_expression_formula(
                formula_text,
                {
                    "start_minutes": start_minutes,
                    "stop_minutes": stop_minutes,
                    "allow_overnight_downtime": bool(settings.get("allow_overnight_downtime", True)),
                    "day_minutes": 24 * 60,
                    "invalid_value": -1,
                },
                default=-1,
            )
            try:
                normalized = int(float(result))
            except Exception:
                normalized = -1
            if normalized >= 0:
                return normalized
            if fallback_label is not None:
                return self.parse_total_minutes(fallback_label)
            return None

        return self.calculate_duration_minutes(
            start_value,
            stop_value,
            allow_overnight=bool(settings.get("allow_overnight_downtime", True)),
        )

    def calculate_stop_time(self, start_value, total_minutes_value, calculation_settings=None):
        settings = self._normalize_calculation_settings(calculation_settings)
        start_text = self.normalize_time_value(start_value)
        total_minutes = self.parse_total_minutes(total_minutes_value)
        if len(start_text) != 4 or total_minutes is None:
            return ""
        start_minutes = int(start_text[:2]) * 60 + int(start_text[2:])
        stop_minutes = (start_minutes + total_minutes) % (24 * 60)

        formulas = settings.get("formulas", {}) if isinstance(settings, dict) else {}
        formula_text = str(formulas.get("downtime_stop_clock") or "").strip()
        if formula_text:
            result = self.evaluate_expression_formula(
                formula_text,
                {
                    "start_minutes": start_minutes,
                    "duration_minutes": total_minutes,
                    "default_stop_minutes": stop_minutes,
                    "day_minutes": 24 * 60,
                },
                default=self._format_compact_time(stop_minutes),
            )
            return self._normalize_clock_formula_result(result)

        return self._format_compact_time(stop_minutes)

    def get_export_directory(self, raw_date):
        base_dir = str(self.settings.get("export_directory", "exports") or "exports").strip() or "exports"
        target_dir = base_dir if os.path.isabs(base_dir) else external_path(base_dir)
        if self.settings.get("organize_exports_by_date", True):
            export_date = self.parse_export_date(raw_date)
            if export_date is not None:
                year_dir = os.path.join(target_dir, export_date.strftime("%Y"))
                legacy_month_dir = os.path.join(year_dir, export_date.strftime("%m"))
                month_folder = export_date.strftime("%m %B")
                target_dir = os.path.join(year_dir, month_folder)
                if os.path.isdir(legacy_month_dir) and not os.path.exists(target_dir):
                    os.makedirs(year_dir, exist_ok=True)
                    os.rename(legacy_month_dir, target_dir)
        os.makedirs(target_dir, exist_ok=True)
        return target_dir

    def _evaluate_safe_expression(self, expression, names=None, functions=None, default=None):
        try:
            return SAFE_EXPRESSION_EVALUATOR.evaluate(expression, names=names, functions=functions)
        except Exception:
            return default

    def _build_formula_functions(self):
        def if_value(condition, true_value, false_value):
            return true_value if condition else false_value

        def float_value(value, fallback=0.0):
            try:
                return float(value)
            except Exception:
                return fallback

        def int_value(value, fallback=0):
            try:
                return int(float(value))
            except Exception:
                return fallback

        def format_clock(value, fallback=""):
            normalized = self._normalize_clock_formula_result(value)
            if normalized:
                return normalized
            return self._normalize_clock_formula_result(fallback)

        return {
            "if_value": if_value,
            "round_minutes": self._round_minutes_value,
            "max_value": max,
            "min_value": min,
            "abs_value": abs,
            "float_value": float_value,
            "int_value": int_value,
            "format_clock": format_clock,
        }

    def calculate_formula(self, formula_str, data_context):
        expression = str(formula_str or "")
        for key, value in (data_context or {}).items():
            placeholder = "{" + key + "}"
            if placeholder in expression:
                normalized_value = value if value not in (None, "") else "0"
                expression = expression.replace(placeholder, str(normalized_value))
        return self._evaluate_safe_expression(expression, default=0)

    def evaluate_expression_formula(self, formula_text, data_context=None, default=None):
        expression = str(formula_text or "").strip()
        if not expression:
            return default
        safe_names = dict(data_context) if isinstance(data_context, dict) else {}
        return self._evaluate_safe_expression(
            expression,
            names=safe_names,
            functions=self._build_formula_functions(),
            default=default,
        )

    def get_production_column(self, column_map, key, fallback=None):
        value = column_map.get(key, fallback)
        text = str(value or "").strip()
        return text or None

    def normalize_header_label(self, value):
        return " ".join(str(value or "").strip().lower().split())

    def detect_production_columns(self, worksheet, configured_columns, start_row):
        resolved_columns = dict(configured_columns)
        header_row = max(1, int(start_row) - 1)
        for column_index in range(1, worksheet.max_column + 1):
            raw_value = worksheet.cell(row=header_row, column=column_index).value
            normalized = self.normalize_header_label(raw_value)
            role_name = PRODUCTION_IMPORT_LABEL_ROLES.get(normalized)
            field_id = self.get_section_field_id_by_role("production", role_name) if role_name else None
            if field_id:
                resolved_columns[field_id] = openpyxl.utils.get_column_letter(column_index)
        return resolved_columns

    def get_row_mapping_config(self, section_name):
        section_config = ROW_SECTION_CONFIG.get(section_name, {})
        mapping_key = section_config.get("mapping_key")
        mapping = self.config.get(mapping_key, {}) if mapping_key else {}
        raw_columns = mapping.get("columns", {}) if isinstance(mapping, dict) else {}
        normalized_columns = {}
        if isinstance(raw_columns, dict):
            for field_id, raw_value in raw_columns.items():
                normalized = self.normalize_row_mapping_column(section_name, field_id, raw_value)
                if normalized.get("column"):
                    normalized_columns[field_id] = normalized
        return {
            "start_row": max(1, int(mapping.get("start_row", 1) or 1)),
            "max_rows": max(1, int(mapping.get("max_rows", section_config.get("default_max_rows", 25)) or 1)),
            "columns": normalized_columns,
        }

    def normalize_row_mapping_column(self, section_name, field_id, raw_value):
        default_export_transform = self.get_default_row_mapping_transform(section_name, field_id, "export")
        default_import_transform = self.get_default_row_mapping_transform(section_name, field_id, "import")
        if isinstance(raw_value, dict):
            return {
                "column": str(raw_value.get("column", "")).strip(),
                "export_enabled": bool(raw_value.get("export_enabled", True)),
                "import_enabled": bool(raw_value.get("import_enabled", True)),
                "export_transform": str(raw_value.get("export_transform", default_export_transform) or default_export_transform).strip(),
                "import_transform": str(raw_value.get("import_transform", default_import_transform) or default_import_transform).strip(),
            }
        return {
            "column": str(raw_value or "").strip(),
            "export_enabled": True,
            "import_enabled": True,
            "export_transform": default_export_transform,
            "import_transform": default_import_transform,
        }

    def get_default_row_mapping_transform(self, section_name, field_id, direction):
        field_role = self.get_section_field_role(section_name, field_id)
        if section_name == "downtime" and field_role == "downtime_code":
            return "code_number" if direction == "export" else "code_lookup"
        if section_name == "downtime" and field_role == "stop_clock":
            return "duration_minutes" if direction == "export" else "stop_from_duration"
        return "value"

    def row_has_values(self, section_name, row_data, field_ids=None):
        candidate_fields = list(field_ids or self.get_open_row_field_ids(section_name) or row_data.keys())
        for field_id in candidate_fields:
            value = row_data.get(field_id)
            if isinstance(value, bool):
                if value:
                    return True
                continue
            if str(value or "").strip():
                return True
        return False

    def resolve_row_export_value(self, section_name, field_id, row_data, column_config, calculation_settings=None):
        value = row_data.get(field_id)
        transform = column_config.get("export_transform", "value")
        if transform == "duration_minutes":
            return self.calculate_runtime_downtime_minutes(
                self.get_row_value_by_role(section_name, row_data, "start_clock", fallback_id="start"),
                self.get_row_value_by_role(section_name, row_data, "stop_clock", fallback_id="stop"),
                calculation_settings=calculation_settings,
                fallback_label=self.get_row_value_by_role(section_name, row_data, "duration_minutes", fallback_id="time_calc"),
            )
        if transform == "code_number":
            return get_code_number(value)
        if transform == "bool_int":
            return 1 if bool(value) else 0
        if transform == "minutes_label":
            return self.parse_total_minutes(value)
        return value

    def resolve_row_import_value(self, section_name, field_id, raw_value, partial_row, column_config, calculation_settings=None):
        transform = column_config.get("import_transform", "value")
        if transform == "code_lookup":
            return normalize_code_value(raw_value)
        if transform == "stop_from_duration":
            return self.calculate_stop_time(
                self.get_row_value_by_role(section_name, partial_row, "start_clock", fallback_id="start"),
                raw_value,
                calculation_settings=calculation_settings,
            )

        field_config = self.get_section_field_config(section_name, field_id)
        if field_config.get("widget") == "checkbutton":
            return self._coerce_bool(raw_value)
        return self.format_cell_value(raw_value)

    def _coerce_bool(self, value):
        if isinstance(value, bool):
            return value
        text = str(value or "").strip().lower()
        if text in {"1", "true", "yes", "on"}:
            return True
        if text in {"0", "false", "no", "off"}:
            return False
        return False

    def _resolve_export_template_path(self):
        template_path = str(self.config.get("template_path", "") or "").strip()
        if not template_path:
            return None
        resolved_path = resource_path(template_path)
        if os.path.exists(resolved_path):
            return resolved_path
        return None

    def _create_export_workbook(self, target_path):
        template_path = self._resolve_export_template_path()
        if template_path:
            shutil.copy(template_path, target_path)
            workbook = openpyxl.load_workbook(target_path)
        else:
            workbook = openpyxl.Workbook()
            workbook.active.title = "Production Log"
        return workbook

    def export_to_template(self, ui_data, shift, date_str, calculation_settings=None):
        clean_date = date_str.replace("/", "")
        export_prefix = str(self.settings.get("default_export_prefix", "Disamatic Production Sheet") or "Disamatic Production Sheet").strip()
        filename = f"{export_prefix} {shift}{clean_date}.xlsx"
        target_path = os.path.join(self.get_export_directory(date_str), filename)
        wb = self._create_export_workbook(target_path)
        ws = wb.active
        for field in self.config["header_fields"]:
            cell_coord = field.get("cell")
            val = ui_data["header"].get(field["id"])
            if cell_coord and field.get("export_enabled", True):
                cell = ws[cell_coord]
                from openpyxl.cell.cell import MergedCell
                if isinstance(cell, MergedCell):
                    for range_ in ws.merged_cells.ranges:
                        if cell_coord in range_:
                            ws.cell(range_.min_row, range_.min_col).value = val
                            break
                else:
                    cell.value = val

        for section_name in ("production", "downtime"):
            mapping = self.get_row_mapping_config(section_name)
            start_row = mapping["start_row"]
            columns = mapping["columns"]
            export_rows = [row for row in ui_data.get(section_name, []) if self.row_has_values(section_name, row)]
            for index, row_data in enumerate(export_rows):
                current_row = start_row + index
                for field_id, column_config in columns.items():
                    if not column_config.get("export_enabled"):
                        continue
                    column_letter = column_config.get("column")
                    if not column_letter:
                        continue
                    ws[f"{column_letter}{current_row}"] = self.resolve_row_export_value(
                        section_name,
                        field_id,
                        row_data,
                        column_config,
                        calculation_settings=calculation_settings,
                    )
        wb.save(target_path)
        wb.close()
        return target_path

    def import_from_excel(self, file_path, calculation_settings=None):
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        formula_workbook = openpyxl.load_workbook(file_path, data_only=False)
        worksheet = workbook.active
        formula_worksheet = formula_workbook.active
        formula_cache = {}
        data = {"header": {}, "production": [], "downtime": []}
        for field in self.config["header_fields"]:
            cell_coord = field.get("cell")
            if cell_coord and field.get("import_enabled", True):
                value = worksheet[cell_coord].value
                if value is None:
                    value = self.resolve_import_cell_value(formula_workbook, formula_worksheet, cell_coord, formula_cache)
                number_format = formula_worksheet[cell_coord].number_format if cell_coord else None
                data["header"][field["id"]] = self.format_header_value(field["id"], value, number_format)
        cast_date_field_id = self.get_header_field_id_by_role("cast_date", fallback_id="cast_date")
        date_field_id = self.get_header_field_id_by_role("log_date", fallback_id="date")
        if cast_date_field_id and not data["header"].get(cast_date_field_id):
            data["header"][cast_date_field_id] = self.compute_cast_date(data["header"].get(date_field_id))

        production_mapping = self.get_row_mapping_config("production")
        detected_columns = self.detect_production_columns(
            formula_worksheet,
            {field_id: config["column"] for field_id, config in production_mapping["columns"].items()},
            production_mapping["start_row"],
        )
        for field_id, column_config in production_mapping["columns"].items():
            if field_id in detected_columns:
                column_config["column"] = detected_columns[field_id]

        for section_name in ("production", "downtime"):
            mapping = production_mapping if section_name == "production" else self.get_row_mapping_config("downtime")
            for index in range(mapping["max_rows"]):
                row_idx = mapping["start_row"] + index
                row_data = {}
                for field_id, column_config in mapping["columns"].items():
                    if not column_config.get("import_enabled"):
                        continue
                    column_letter = column_config.get("column")
                    if not column_letter:
                        continue
                    cell_ref = f"{column_letter}{row_idx}"
                    raw_value = worksheet[cell_ref].value
                    if raw_value is None:
                        raw_value = self.resolve_import_cell_value(formula_workbook, formula_worksheet, cell_ref, formula_cache)
                    row_data[field_id] = self.resolve_row_import_value(
                        section_name,
                        field_id,
                        raw_value,
                        row_data,
                        column_config,
                        calculation_settings=calculation_settings,
                    )

                if not self.row_has_values(section_name, row_data, mapping["columns"].keys()):
                    break
                data[section_name].append(row_data)

        workbook.close()
        formula_workbook.close()
        return data
