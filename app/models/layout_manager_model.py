# Production Logging Center (GLC Edition)
# Copyright (C) 2026 Jamie Martin
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
import json
import os
import re
import shutil
from copy import deepcopy
from datetime import datetime

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from app.layout_config_service import LayoutConfigService
from app.persistence import write_json_with_backup
from app.production_log_roles import PROTECTED_ROW_ROLES, REQUIRED_MAPPING_ROLES, get_default_row_field_id, normalize_role_name, normalize_row_section_name, resolve_header_field_role, resolve_row_field_role
from app.utils import external_path, local_or_resource_path

__module_name__ = "Layout Manager"
__version__ = "1.0.2"

VALID_IMPORT_TRANSFORMS = ("value", "code_lookup", "stop_from_duration")
VALID_EXPORT_TRANSFORMS = ("value", "code_number", "duration_minutes", "bool_int", "minutes_label")
DEFAULT_MAPPING_MAX_ROWS = 25
LAYOUT_VERSION_FILE = "data/config/layout_config_versions.json"
DEFAULT_SECTIONS = (
    {
        "id": "header",
        "name": "Header Fields",
        "description": "Single-record header fields",
        "fields_key": "header_fields",
        "section_type": "single",
        "behavior_profile": "header",
    },
    {
        "id": "production",
        "name": "Production Row Fields",
        "description": "Repeating production rows",
        "fields_key": "production_row_fields",
        "mapping_key": "production_mapping",
        "section_type": "repeating",
        "behavior_profile": "production",
        "default_max_rows": 50,
        "delete_row_policy": {
            "show_delete_button": True,
            "delete_button_label": "X",
            "delete_button_tooltip": "Delete this row",
            "require_delete_confirmation": False,
        },
    },
    {
        "id": "downtime",
        "name": "Downtime Row Fields",
        "description": "Repeating downtime rows",
        "fields_key": "downtime_row_fields",
        "mapping_key": "downtime_mapping",
        "section_type": "repeating",
        "behavior_profile": "downtime",
        "default_max_rows": 25,
        "delete_row_policy": {
            "show_delete_button": True,
            "delete_button_label": "X",
            "delete_button_tooltip": "Delete this row",
            "require_delete_confirmation": False,
        },
    },
)


class LayoutManagerModel:
    REQUIRED_TOP_LEVEL_KEYS = (
        "template_path",
        "header_fields",
        "sections",
        "export_prefix",
    )
    EDITOR_REQUIRED_TOP_LEVEL_KEYS = REQUIRED_TOP_LEVEL_KEYS
    EDITOR_OPTIONAL_TOP_LEVEL_KEYS = ("editor_presets", "calculations")
    EDITOR_TOP_LEVEL_KEYS = EDITOR_REQUIRED_TOP_LEVEL_KEYS + EDITOR_OPTIONAL_TOP_LEVEL_KEYS
    EDITOR_PRESET_SECTION_KEYS = ("header_fields",)

    def __init__(self):
        self.service = LayoutConfigService()
        self.is_dirty = False
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        self._default_config_template = None
        self.protected_field_ids = set()
        self.protected_header_roles = set()
        self.row_field_sections = ("production_row_fields", "downtime_row_fields")
        self.protected_row_field_ids = {
            "production_row_fields": {"shop_order", "part_number", "rate_lookup", "rate_override_enabled", "molds", "time_calc"},
            "downtime_row_fields": {"start", "stop", "code", "cause", "time_calc"},
        }
        self.protected_row_roles = {section_name: set(PROTECTED_ROW_ROLES.get(section_name, set())) for section_name in self.row_field_sections}
        self._import_export_metadata_cache = {}
        self._template_workbook_stats_cache = {}

    @property
    def local_config(self):
        return self.service.local_config

    @property
    def internal_config(self):
        return self.service.internal_config

    @property
    def config_path(self):
        return self.current_source_path

    @config_path.setter
    def config_path(self, value):
        self.current_source_path = value

    @property
    def save_path(self):
        return self.current_save_path

    def get_active_form_info(self):
        return self.service.get_active_form_info()

    def get_form_info(self, form_id):
        return self.service.get_form_info(form_id)

    def list_forms(self):
        return self.service.list_forms()

    def serialize_config(self, config):
        return json.dumps(config, indent=4)

    def parse_editor_text(self, text, base_config=None):
        config, _payload_details = self.resolve_editor_text(text, base_config=base_config)
        return config

    def _get_default_config_template(self):
        if self._default_config_template is None:
            try:
                default_config, _source_path = self.service.load_default()
            except Exception:
                default_config = {}
            normalized_default = dict(default_config) if isinstance(default_config, dict) else {}
            normalized_default["sections"] = self._normalize_sections(normalized_default)
            if "header_fields" in normalized_default:
                normalized_default["header_fields"] = self._normalize_header_fields(
                    deepcopy(normalized_default.get("header_fields"))
                )
            for binding in self._iter_repeating_section_bindings(normalized_default):
                section_name = binding["fields_key"]
                mapping_name = binding["mapping_key"]
                if section_name in normalized_default:
                    normalized_default[section_name] = self._normalize_row_fields(
                        normalized_default.get(section_name),
                        section_name,
                    )
                if mapping_name in normalized_default:
                    normalized_default[mapping_name] = self._normalize_mapping(
                        mapping_name,
                        normalized_default.get(mapping_name),
                        row_fields=normalized_default.get(section_name),
                    )
            normalized_default["template_path"] = str(normalized_default.get("template_path") or "")
            active_info = None
            try:
                active_info = self.service.get_active_form_info()
            except Exception:
                pass
            fallback_prefix = active_info.get("name", "Disamatic Production Sheet") if active_info else "Disamatic Production Sheet"
            if "export_prefix" in normalized_default:
                normalized_default["export_prefix"] = str(normalized_default["export_prefix"] if normalized_default["export_prefix"] is not None else "").strip()
            else:
                normalized_default["export_prefix"] = str(fallback_prefix).strip()
            if "calculations" in normalized_default and isinstance(normalized_default.get("calculations"), dict):
                normalized_default["calculations"] = deepcopy(normalized_default.get("calculations"))
            self._default_config_template = normalized_default
        return deepcopy(self._default_config_template)

    def _normalize_sections(self, config):
        raw_sections = config.get("sections") if isinstance(config, dict) else None
        if not isinstance(raw_sections, list):
            raw_sections = []
        if not raw_sections:
            raw_sections = deepcopy(DEFAULT_SECTIONS)

        normalized_sections = []
        seen_ids = set()

        for raw_section in raw_sections:
            if not isinstance(raw_section, dict):
                continue
            section_id = str(raw_section.get("id", "")).strip().lower()
            if not section_id or section_id in seen_ids:
                continue
            normalized_section = deepcopy(raw_section)
            normalized_section["id"] = section_id
            normalized_section["name"] = str(
                raw_section.get("name", normalized_section.get("name", section_id.replace("_", " ").title()))
            ).strip() or section_id.replace("_", " ").title()

            description_text = str(raw_section.get("description", normalized_section.get("description", ""))).strip()
            if description_text:
                normalized_section["description"] = description_text
            else:
                normalized_section.pop("description", None)

            fields_key = str(raw_section.get("fields_key", normalized_section.get("fields_key", ""))).strip()
            if fields_key:
                normalized_section["fields_key"] = fields_key
            else:
                normalized_section.pop("fields_key", None)

            mapping_key = str(raw_section.get("mapping_key", normalized_section.get("mapping_key", ""))).strip()
            if mapping_key:
                normalized_section["mapping_key"] = mapping_key
            else:
                normalized_section.pop("mapping_key", None)

            section_type = str(raw_section.get("section_type", normalized_section.get("section_type", "single"))).strip().lower()
            normalized_section["section_type"] = section_type or "single"

            behavior_profile = normalize_role_name(
                raw_section.get("behavior_profile", normalized_section.get("behavior_profile", section_id))
            )
            if behavior_profile:
                normalized_section["behavior_profile"] = behavior_profile
            else:
                normalized_section.pop("behavior_profile", None)

            if normalized_section["section_type"] == "repeating":
                default_max_rows = raw_section.get("default_max_rows", normalized_section.get("default_max_rows"))
                if default_max_rows not in (None, ""):
                    try:
                        normalized_section["default_max_rows"] = max(1, int(default_max_rows))
                    except (TypeError, ValueError):
                        normalized_section["default_max_rows"] = default_max_rows
                else:
                    normalized_section.pop("default_max_rows", None)
                default_field_width = raw_section.get("default_field_width", normalized_section.get("default_field_width"))
                if default_field_width not in (None, ""):
                    try:
                        normalized_section["default_field_width"] = max(1, int(default_field_width))
                    except (TypeError, ValueError):
                        normalized_section.pop("default_field_width", None)
                else:
                    normalized_section.pop("default_field_width", None)
                raw_policy = raw_section.get("delete_row_policy", normalized_section.get("delete_row_policy"))
                if isinstance(raw_policy, dict):
                    normalized_section["delete_row_policy"] = self._normalize_delete_row_policy(raw_policy)
                else:
                    normalized_section.pop("delete_row_policy", None)
            else:
                normalized_section.pop("default_max_rows", None)
                normalized_section.pop("default_field_width", None)
                normalized_section.pop("delete_row_policy", None)

            normalized_sections.append(normalized_section)
            seen_ids.add(section_id)

        return normalized_sections

    def _normalize_delete_row_policy(self, policy, default_policy=None):
        base_policy = dict(default_policy or {})
        raw_policy = policy if isinstance(policy, dict) else {}
        show_delete_button = self._normalize_bool_value(
            raw_policy.get("show_delete_button", base_policy.get("show_delete_button", True)),
            default=True,
        )
        delete_button_label = str(
            raw_policy.get("delete_button_label", base_policy.get("delete_button_label", "X")) or "X"
        ).strip() or "X"
        delete_button_tooltip = str(
            raw_policy.get("delete_button_tooltip", base_policy.get("delete_button_tooltip", "Delete this row"))
            or "Delete this row"
        ).strip() or "Delete this row"
        require_delete_confirmation = self._normalize_bool_value(
            raw_policy.get("require_delete_confirmation", base_policy.get("require_delete_confirmation", False)),
            default=False,
        )
        return {
            "show_delete_button": show_delete_button,
            "delete_button_label": delete_button_label,
            "delete_button_tooltip": delete_button_tooltip,
            "require_delete_confirmation": require_delete_confirmation,
        }

    def get_sections(self, config=None):
        config_data = self.normalize_config(config) if isinstance(config, dict) else self._get_default_config_template()
        return [deepcopy(section) for section in config_data.get("sections", []) if isinstance(section, dict)]

    def get_section_info(self, section_id, config=None):
        normalized_section_id = str(section_id or "").strip().lower()
        for section in self.get_sections(config=config):
            if section.get("id") == normalized_section_id:
                return section
        return {}

    def get_section_name(self, section_id, config=None, fallback_name=None):
        section_info = self.get_section_info(section_id, config=config)
        if section_info:
            return section_info.get("name") or fallback_name or str(section_id or "").replace("_", " ").title()
        return fallback_name or str(section_id or "").replace("_", " ").title()

    def build_editor_guardrails(self, config=None):
        config_data = self.normalize_config(config) if isinstance(config, dict) else self._get_default_config_template()
        sections = [section for section in config_data.get("sections", []) if isinstance(section, dict)]
        profile_matches = {profile_name: [] for profile_name in ("header", "production", "downtime")}
        for section in sections:
            profile_name = normalize_role_name(section.get("behavior_profile"))
            if profile_name in profile_matches:
                profile_matches[profile_name].append(section)

        routed_sections = []
        warnings = []
        for profile_name, matches in profile_matches.items():
            if len(matches) == 1:
                section = matches[0]
                routed_sections.append(
                    {
                        "profile": profile_name,
                        "name": section.get("name") or section.get("id", profile_name).replace("_", " ").title(),
                        "section_type": section.get("section_type", "single"),
                        "fields_key": section.get("fields_key", ""),
                        "mapping_key": section.get("mapping_key", ""),
                    }
                )
                continue
            if not matches:
                warnings.append(f"No section currently routes the supported '{profile_name}' profile.")
                continue
            warnings.append(f"Multiple sections claim the supported '{profile_name}' profile. Keep supported profiles unique.")

        notes = [
            "Import and export only move fields that are both listed in the active section schema and enabled in the mapping or header toggle.",
            "Supported routing profiles are bounded to header, production, and downtime. Unsupported profiles fail closed.",
            "Normalization preserves the current JSON shape and validation flags mapping or role issues without silently rebuilding removed sections or fields.",
        ]
        return {
            "routed_sections": routed_sections,
            "warnings": warnings,
            "notes": notes,
        }

    def update_section_metadata(self, config, section_id, section_values):
        normalized_section_id = str(section_id or "").strip().lower()
        if not normalized_section_id:
            raise ValueError("Section ID is missing.")

        config["sections"] = self._normalize_sections(config)
        target_section = None
        for section in config["sections"]:
            if section.get("id") == normalized_section_id:
                target_section = section
                break
        if target_section is None:
            raise ValueError(f"Section '{normalized_section_id}' was not found.")

        name_text = str(section_values.get("name", target_section.get("name", ""))).strip()
        if not name_text:
            raise ValueError("Section name cannot be empty.")
        target_section["name"] = name_text

        description_text = str(section_values.get("description", target_section.get("description", ""))).strip()
        if description_text:
            target_section["description"] = description_text
        else:
            target_section.pop("description", None)

        section_type = str(section_values.get("section_type", target_section.get("section_type", "single"))).strip().lower()
        if section_type not in {"single", "repeating"}:
            raise ValueError("Section type must be 'single' or 'repeating'.")
        target_section["section_type"] = section_type

        behavior_profile = str(section_values.get("behavior_profile", target_section.get("behavior_profile", normalized_section_id))).strip().lower()
        if not behavior_profile:
            raise ValueError("Behavior profile cannot be empty.")
        target_section["behavior_profile"] = behavior_profile

        if section_type == "repeating":
            default_max_rows_text = str(
                section_values.get("default_max_rows", target_section.get("default_max_rows", DEFAULT_MAPPING_MAX_ROWS))
            ).strip()
            try:
                target_section["default_max_rows"] = max(1, int(default_max_rows_text or DEFAULT_MAPPING_MAX_ROWS))
            except (TypeError, ValueError):
                raise ValueError("Default max rows must be a positive integer.")

            default_field_width_text = str(
                section_values.get("default_field_width", target_section.get("default_field_width", ""))
            ).strip()
            if default_field_width_text:
                try:
                    target_section["default_field_width"] = max(1, int(default_field_width_text))
                except (TypeError, ValueError):
                    raise ValueError("Default field width must be a positive integer.")
            else:
                target_section.pop("default_field_width", None)

            target_section["delete_row_policy"] = self._normalize_delete_row_policy(
                {
                    "show_delete_button": section_values.get(
                        "show_delete_button",
                        (target_section.get("delete_row_policy") or {}).get("show_delete_button", True),
                    ),
                    "delete_button_label": section_values.get(
                        "delete_button_label",
                        (target_section.get("delete_row_policy") or {}).get("delete_button_label", "X"),
                    ),
                    "delete_button_tooltip": section_values.get(
                        "delete_button_tooltip",
                        (target_section.get("delete_row_policy") or {}).get("delete_button_tooltip", "Delete this row"),
                    ),
                    "require_delete_confirmation": section_values.get(
                        "require_delete_confirmation",
                        (target_section.get("delete_row_policy") or {}).get("require_delete_confirmation", False),
                    ),
                },
                default_policy=target_section.get("delete_row_policy"),
            )
        else:
            target_section.pop("default_max_rows", None)
            target_section.pop("default_field_width", None)
            target_section.pop("delete_row_policy", None)

        return config, f"Updated section '{normalized_section_id}' metadata"

    def move_section(self, config, section_id, direction):
        normalized_section_id = str(section_id or "").strip().lower()
        sections = config.get("sections") if isinstance(config.get("sections"), list) else []
        index = next((position for position, section in enumerate(sections) if section.get("id") == normalized_section_id), None)
        if index is None:
            raise ValueError(f"Section '{normalized_section_id}' was not found.")
        target_index = index + int(direction)
        if target_index < 0 or target_index >= len(sections):
            return config, None
        sections[index], sections[target_index] = sections[target_index], sections[index]
        config["sections"] = sections
        return config, f"Reordered section '{normalized_section_id}'"

    def add_section(self, config, section_values):
        section_id = normalize_role_name(section_values.get("id"))
        if not section_id:
            raise ValueError("Section ID is required.")
        sections = config.get("sections") if isinstance(config.get("sections"), list) else []
        if any(str(section.get("id", "")).strip().lower() == section_id for section in sections):
            raise ValueError(f"Section '{section_id}' already exists.")

        section_name = str(section_values.get("name") or "").strip() or section_id.replace("_", " ").title()
        section_description = str(section_values.get("description") or "").strip()
        section_type = str(section_values.get("section_type") or "single").strip().lower()
        if section_type not in {"single", "repeating"}:
            raise ValueError("Section type must be 'single' or 'repeating'.")
        behavior_profile = normalize_role_name(section_values.get("behavior_profile")) or section_id

        if section_id == "header":
            fields_key = "header_fields"
        elif section_type == "repeating":
            fields_key = f"{section_id}_row_fields"
        else:
            fields_key = f"{section_id}_fields"

        section_payload = {
            "id": section_id,
            "name": section_name,
            "fields_key": fields_key,
            "section_type": section_type,
            "behavior_profile": behavior_profile,
        }
        if section_description:
            section_payload["description"] = section_description

        if fields_key not in config or not isinstance(config.get(fields_key), list):
            config[fields_key] = []

        if section_type == "repeating":
            max_rows_text = str(section_values.get("default_max_rows") or DEFAULT_MAPPING_MAX_ROWS).strip()
            try:
                section_payload["default_max_rows"] = max(1, int(max_rows_text or DEFAULT_MAPPING_MAX_ROWS))
            except (TypeError, ValueError):
                raise ValueError("Default max rows must be a positive integer.")
            field_width_text = str(section_values.get("default_field_width") or "").strip()
            if field_width_text:
                try:
                    section_payload["default_field_width"] = max(1, int(field_width_text))
                except (TypeError, ValueError):
                    raise ValueError("Default field width must be a positive integer.")
            mapping_key = f"{section_id}_mapping"
            section_payload["mapping_key"] = mapping_key
            if mapping_key not in config or not isinstance(config.get(mapping_key), dict):
                config[mapping_key] = {
                    "start_row": 1,
                    "max_rows": section_payload["default_max_rows"],
                    "columns": {},
                }
            section_payload["delete_row_policy"] = self._normalize_delete_row_policy(
                {
                    "show_delete_button": section_values.get("show_delete_button", True),
                    "delete_button_label": section_values.get("delete_button_label", "X"),
                    "delete_button_tooltip": section_values.get("delete_button_tooltip", "Delete this row"),
                    "require_delete_confirmation": section_values.get("require_delete_confirmation", False),
                }
            )

        sections.append(section_payload)
        config["sections"] = self._normalize_sections({"sections": sections})
        return config, f"Added section '{section_id}'"

    def remove_section(self, config, section_id):
        normalized_section_id = str(section_id or "").strip().lower()
        sections = config.get("sections") if isinstance(config.get("sections"), list) else []
        target_section = next((section for section in sections if section.get("id") == normalized_section_id), None)
        if target_section is None:
            raise ValueError(f"Section '{normalized_section_id}' was not found.")

        fields_key = str(target_section.get("fields_key") or "").strip()
        mapping_key = str(target_section.get("mapping_key") or "").strip()
        config["sections"] = [section for section in sections if section.get("id") != normalized_section_id]
        if fields_key and fields_key in config and fields_key != "header_fields":
            config.pop(fields_key, None)
        if mapping_key and mapping_key in config:
            config.pop(mapping_key, None)
        return config, f"Removed section '{normalized_section_id}'"

    def _get_mapping_section_name(self, mapping_name, config=None):
        normalized_mapping_name = str(mapping_name or "").strip()
        config_data = config if isinstance(config, dict) else {}
        if config_data:
            for binding in self._iter_repeating_section_bindings(config_data):
                if binding.get("mapping_key") == normalized_mapping_name:
                    return normalize_row_section_name(binding.get("section_id") or binding.get("fields_key") or "")
        if normalized_mapping_name.endswith("_mapping"):
            return normalize_row_section_name(normalized_mapping_name[:-8])
        return ""

    def _get_default_mapping_transform(self, mapping_name, field_id, direction, row_fields=None):
        section_name = self._get_mapping_section_name(mapping_name)
        role_name = ""
        for field in row_fields if isinstance(row_fields, list) else []:
            if not isinstance(field, dict):
                continue
            if str(field.get("id", "")).strip() == str(field_id or "").strip():
                role_name = resolve_row_field_role(section_name, field_id, field.get("role"))
                break
        if section_name == "downtime" and role_name == "downtime_code":
            return "code_number" if direction == "export" else "code_lookup"
        if section_name == "downtime" and role_name == "stop_clock":
            return "duration_minutes" if direction == "export" else "stop_from_duration"
        return "value"

    def _normalize_mapping_column_config(self, mapping_name, field_id, raw_value, row_fields=None):
        default_export_transform = self._get_default_mapping_transform(mapping_name, field_id, "export", row_fields=row_fields)
        default_import_transform = self._get_default_mapping_transform(mapping_name, field_id, "import", row_fields=row_fields)
        if isinstance(raw_value, dict):
            return {
                "column": str(raw_value.get("column", "")).strip(),
                "import_enabled": self._normalize_bool_value(raw_value.get("import_enabled"), default=True),
                "export_enabled": self._normalize_bool_value(raw_value.get("export_enabled"), default=True),
                "import_transform": str(raw_value.get("import_transform", default_import_transform) or default_import_transform).strip() or default_import_transform,
                "export_transform": str(raw_value.get("export_transform", default_export_transform) or default_export_transform).strip() or default_export_transform,
            }
        return str(raw_value or "").strip()

    def _build_default_mapping(self, mapping_name, row_fields):
        return {
            "start_row": 1,
            "max_rows": DEFAULT_MAPPING_MAX_ROWS,
            "columns": {
                field["id"]: self._normalize_mapping_column_config(mapping_name, field["id"], "", row_fields=row_fields)
                for field in row_fields
                if isinstance(field, dict) and field.get("id")
            },
        }

    def _normalize_header_field_entry(self, field):
        if not isinstance(field, dict):
            return None
        normalized_field = deepcopy(field)
        normalized_field.pop("_original_id", None)
        field_id = str(normalized_field.get("id", "")).strip()
        role_name = resolve_header_field_role(field_id, normalized_field.get("role"))
        if role_name:
            normalized_field["role"] = role_name
        else:
            normalized_field.pop("role", None)

        widget_name = str(normalized_field.get("widget", "entry") or "entry").strip().lower() or "entry"
        if widget_name not in {"entry", "combobox"}:
            widget_name = "entry"
        if widget_name == "entry" and "widget" not in field:
            normalized_field.pop("widget", None)
        else:
            normalized_field["widget"] = widget_name

        normalized_field["import_enabled"] = self._normalize_bool_value(
            normalized_field.get("import_enabled"),
            default=True,
        )
        normalized_field["export_enabled"] = self._normalize_bool_value(
            normalized_field.get("export_enabled"),
            default=True,
        )

        state_name = str(normalized_field.get("state", "") or "").strip().lower()
        if state_name in {"normal", "disabled", "readonly"}:
            normalized_field["state"] = state_name
        else:
            normalized_field.pop("state", None)

        options_source_name = str(normalized_field.get("options_source", "") or "").strip().lower()
        if widget_name == "combobox":
            from app.downtime_codes import get_available_options_sources
            if options_source_name in get_available_options_sources():
                normalized_field["options_source"] = options_source_name
            else:
                normalized_field.pop("options_source", None)
            self._set_optional_list_field(normalized_field, "values", normalized_field.get("values"))
        else:
            normalized_field.pop("options_source", None)
            normalized_field.pop("values", None)

        readonly_enabled = self._normalize_bool_value(normalized_field.get("readonly"), default=False)
        if field_id == "cast_date":
            normalized_field["readonly"] = True
            normalized_field.pop("default", None)
        elif readonly_enabled:
            normalized_field["readonly"] = True
        else:
            normalized_field.pop("readonly", None)

        return normalized_field

    def _normalize_header_fields(self, header_fields):
        if not isinstance(header_fields, list):
            return []

        normalized_fields = []
        for field in header_fields:
            normalized_field = self._normalize_header_field_entry(field)
            if not isinstance(normalized_field, dict):
                continue
            normalized_fields.append(normalized_field)
        return normalized_fields

    def _normalize_row_field_entry(self, field, section_name):
        if not isinstance(field, dict):
            return None
        normalized_field = deepcopy(field)
        field_id = str(normalized_field.get("id", "")).strip()
        widget_name = str(normalized_field.get("widget", "entry") or "entry").strip().lower() or "entry"
        normalized_field["widget"] = widget_name
        role_name = resolve_row_field_role(section_name, field_id, normalized_field.get("role"))
        if role_name:
            normalized_field["role"] = role_name
        else:
            normalized_field.pop("role", None)
        if widget_name == "combobox":
            self._set_optional_list_field(normalized_field, "values", normalized_field.get("values"))
        else:
            normalized_field.pop("values", None)
        normalized_field.pop("_original_id", None)

        for key in ("readonly", "derived", "math_trigger", "open_row_trigger", "user_input", "expand", "bold"):
            if key in normalized_field:
                val = self._normalize_bool_value(normalized_field.get(key), default=False)
                if val:
                    normalized_field[key] = True
                else:
                    normalized_field.pop(key, None)

        if normalized_field.get("user_input"):
            normalized_field.pop("derived", None)

        return normalized_field

    def _normalize_row_fields(self, row_fields, section_name):
        if not isinstance(row_fields, list):
            return []

        normalized_fields = []
        seen_ids = set()
        for field in row_fields:
            normalized_field = self._normalize_row_field_entry(field, section_name)
            if not isinstance(normalized_field, dict):
                continue
            field_id = str(normalized_field.get("id", "")).strip()
            if not field_id or field_id in seen_ids:
                continue
            normalized_fields.append(normalized_field)
            seen_ids.add(field_id)

        return normalized_fields

    def _normalize_editor_presets(self, editor_presets):
        if not isinstance(editor_presets, dict):
            return {}

        config_data = self._get_default_config_template()
        row_section_names = [binding["fields_key"] for binding in self._iter_repeating_section_bindings(config_data)]
        if not row_section_names:
            row_section_names = list(self.row_field_sections)

        normalized_presets = {}
        raw_header_presets = editor_presets.get("header_fields")
        if isinstance(raw_header_presets, list):
            normalized_presets["header_fields"] = self._normalize_header_fields(deepcopy(raw_header_presets))

        for section_name in row_section_names:
            raw_row_presets = editor_presets.get(section_name)
            if not isinstance(raw_row_presets, list):
                continue
            normalized_presets[section_name] = [
                normalized_field
                for normalized_field in (
                    self._normalize_row_field_entry(field, section_name)
                    for field in raw_row_presets
                )
                if isinstance(normalized_field, dict)
            ]

        return normalized_presets

    def _merge_row_fields_with_defaults(self, row_fields, default_row_fields=None, section_name=""):
        return self._normalize_row_fields(row_fields, section_name)

    def _normalize_mapping(self, mapping_name, mapping, row_fields=None):
        if not isinstance(mapping, dict):
            return {}
        normalized_mapping = {}
        for key_name, value in mapping.items():
            if key_name == "columns" and isinstance(value, dict):
                normalized_columns = {}
                for field_id, raw_column_value in value.items():
                    normalized_field_id = str(field_id or "").strip()
                    if not normalized_field_id:
                        continue
                    normalized_columns[normalized_field_id] = self._normalize_mapping_column_config(
                        mapping_name,
                        normalized_field_id,
                        raw_column_value,
                        row_fields=row_fields,
                    )
                normalized_mapping["columns"] = normalized_columns
            else:
                normalized_mapping[key_name] = deepcopy(value)
        return normalized_mapping

    def _normalize_calculation_section_profiles(self, raw_profiles):
        if not isinstance(raw_profiles, list):
            return []
        normalized_profiles = []
        seen_sections = set()
        for raw_profile in raw_profiles:
            if not isinstance(raw_profile, dict):
                continue
            section_id = normalize_role_name(raw_profile.get("section_id"))
            if not section_id or section_id in seen_sections:
                continue
            requires_calculations = self._normalize_bool_value(raw_profile.get("requires_calculations"), default=True)
            calculation_profile = normalize_role_name(raw_profile.get("calculation_profile")) or section_id
            normalized_profiles.append(
                {
                    "section_id": section_id,
                    "requires_calculations": requires_calculations,
                    "calculation_profile": calculation_profile,
                }
            )
            seen_sections.add(section_id)
        return normalized_profiles

    def _build_default_calculation_section_profiles(self, config):
        profiles = []
        config_data = config if isinstance(config, dict) else {}
        sections = config_data.get("sections") if isinstance(config_data.get("sections"), list) else []
        for section in sections:
            if not isinstance(section, dict):
                continue
            section_type = str(section.get("section_type") or "single").strip().lower()
            if section_type != "repeating":
                continue
            section_id = normalize_role_name(section.get("id"))
            if not section_id:
                continue
            behavior_profile = normalize_role_name(section.get("behavior_profile")) or section_id
            requires_calculations = behavior_profile in {"production", "downtime"}
            profiles.append(
                {
                    "section_id": section_id,
                    "requires_calculations": requires_calculations,
                    "calculation_profile": behavior_profile,
                }
            )
        return profiles

    def _normalize_calculations_metadata(self, config, form_id=None):
        config_data = config if isinstance(config, dict) else {}
        raw_calculations = config_data.get("calculations") if isinstance(config_data.get("calculations"), dict) else {}
        resolved_form_id = str(form_id or "").strip().lower()
        if not resolved_form_id:
            resolved_form_id = "form"
        default_companion_path = os.path.join("data", "forms", resolved_form_id, "calculations.json").replace("\\", "/")
        companion_relative_path = str(raw_calculations.get("companion_relative_path") or default_companion_path).strip().replace("\\", "/")
        section_profiles = self._normalize_calculation_section_profiles(raw_calculations.get("section_profiles"))
        if not section_profiles:
            section_profiles = self._build_default_calculation_section_profiles(config_data)
        return {
            "companion_relative_path": companion_relative_path,
            "section_profiles": section_profiles,
        }

    def _iter_repeating_section_bindings(self, config):
        sections = config.get("sections") if isinstance(config, dict) and isinstance(config.get("sections"), list) else []
        bindings = []
        seen_fields_keys = set()

        for section in sections:
            if not isinstance(section, dict):
                continue
            if str(section.get("section_type") or "single").strip().lower() != "repeating":
                continue
            fields_key = str(section.get("fields_key") or "").strip()
            mapping_key = str(section.get("mapping_key") or "").strip()
            if not fields_key or not mapping_key or fields_key in seen_fields_keys:
                continue
            bindings.append(
                {
                    "section_id": str(section.get("id") or "").strip(),
                    "behavior_profile": str(section.get("behavior_profile") or "").strip().lower(),
                    "fields_key": fields_key,
                    "mapping_key": mapping_key,
                }
            )
            seen_fields_keys.add(fields_key)

        if bindings:
            return bindings

        config_data = config if isinstance(config, dict) else {}
        inferred_bindings = []
        for key_name, value in config_data.items():
            normalized_key = str(key_name or "").strip()
            if not normalized_key.endswith("_row_fields") or not isinstance(value, list):
                continue
            mapping_key = f"{normalized_key[:-11]}_mapping"
            if not isinstance(config_data.get(mapping_key), dict):
                continue
            inferred_bindings.append(
                {
                    "section_id": normalize_row_section_name(normalized_key),
                    "behavior_profile": normalize_row_section_name(normalized_key),
                    "fields_key": normalized_key,
                    "mapping_key": mapping_key,
                }
            )

        if inferred_bindings:
            return inferred_bindings

        return [
            {
                "section_id": normalize_row_section_name(section.get("id") or section.get("fields_key") or ""),
                "behavior_profile": normalize_role_name(section.get("behavior_profile") or section.get("id") or ""),
                "fields_key": section["fields_key"],
                "mapping_key": section["mapping_key"],
            }
            for section in DEFAULT_SECTIONS
            if isinstance(section, dict)
            and str(section.get("section_type") or "").strip().lower() == "repeating"
            and str(section.get("fields_key") or "").strip()
            and str(section.get("mapping_key") or "").strip()
        ]

    def normalize_config(self, config, form_info=None):
        if not isinstance(config, dict):
            return self._get_default_config_template()

        normalized = deepcopy(config)
        normalized["sections"] = self._normalize_sections(normalized)
        if "header_fields" in normalized:
            normalized["header_fields"] = self._normalize_header_fields(
                deepcopy(normalized.get("header_fields"))
            )
        repeating_bindings = self._iter_repeating_section_bindings(normalized)
        for binding in repeating_bindings:
            section_name = binding["fields_key"]
            mapping_name = binding["mapping_key"]
            if section_name in normalized:
                normalized[section_name] = self._normalize_row_fields(
                    normalized.get(section_name),
                    section_name,
                )
            if mapping_name in normalized:
                normalized[mapping_name] = self._normalize_mapping(
                    mapping_name,
                    normalized.get(mapping_name),
                    row_fields=normalized.get(section_name),
                )
        if "editor_presets" in normalized:
            normalized["editor_presets"] = self._normalize_editor_presets(normalized.get("editor_presets"))
        normalized["calculations"] = self._normalize_calculations_metadata(normalized)
        if "template_path" in normalized:
            normalized["template_path"] = str(normalized.get("template_path") or "")

        form_name = None
        if isinstance(form_info, dict):
            form_name = form_info.get("name")
        if not form_name:
            try:
                active_info = self.service.get_active_form_info()
                form_name = active_info.get("name")
            except Exception:
                pass
        if not form_name:
            form_name = "Disamatic Production Sheet"

        if "export_prefix" in normalized:
            normalized["export_prefix"] = str(normalized["export_prefix"] if normalized["export_prefix"] is not None else "").strip()
        else:
            normalized["export_prefix"] = str(form_name).strip()
        return normalized

    def migrate_forms_to_scoped_storage(self):
        registry = self.service.registry
        registry_payload = registry.get_registry()
        forms = registry_payload.get("forms") if isinstance(registry_payload.get("forms"), list) else []
        migrated = []
        skipped = []

        legacy_calculations_path = self.service.data_registry.resolve_read_path("production_log_calculations")
        legacy_calculation_payload = {}
        if os.path.exists(legacy_calculations_path):
            try:
                with open(legacy_calculations_path, "r", encoding="utf-8") as handle:
                    loaded_payload = json.load(handle)
                if isinstance(loaded_payload, dict):
                    legacy_calculation_payload = loaded_payload
            except Exception:
                legacy_calculation_payload = {}

        for index, form_record in enumerate(forms):
            if not isinstance(form_record, dict):
                continue
            form_id = str(form_record.get("id") or "").strip().lower()
            if not form_id:
                skipped.append(f"index:{index} (Missing ID)")
                continue
            if form_record.get("built_in"):
                skipped.append(f"{form_id} (Built-in default)")
                continue

            form_info = registry.enrich_form_record(form_record, active_form_id=registry_payload.get("active_form_id"))
            target_layout_relative_path = os.path.join("data", "forms", form_id, "layout.json").replace("\\", "/")
            target_calculation_relative_path = os.path.join("data", "forms", form_id, "calculations.json").replace("\\", "/")
            target_layout_path = external_path(target_layout_relative_path)
            target_calculation_path = external_path(target_calculation_relative_path)
            os.makedirs(os.path.dirname(target_layout_path), exist_ok=True)
            os.makedirs(os.path.dirname(target_calculation_path), exist_ok=True)

            try:
                with open(form_info.get("load_path"), "r", encoding="utf-8") as handle:
                    config_payload = json.load(handle)
            except Exception as exc:
                skipped.append(f"{form_id} (Load error: {exc})")
                continue

            normalized_config = self.normalize_config(config_payload, form_info=form_info)
            normalized_config["calculations"] = self._normalize_calculations_metadata(normalized_config, form_id=form_id)
            normalized_config["calculations"]["companion_relative_path"] = target_calculation_relative_path

            old_calculations_relative_path = str((config_payload.get("calculations") or {}).get("companion_relative_path") or "").strip().replace("\\", "/")
            old_calculation_payload = {}
            if old_calculations_relative_path:
                old_calculation_path = external_path(old_calculations_relative_path)
                if os.path.exists(old_calculation_path):
                    try:
                        with open(old_calculation_path, "r", encoding="utf-8") as handle:
                            loaded_payload = json.load(handle)
                        if isinstance(loaded_payload, dict):
                            old_calculation_payload = loaded_payload
                    except Exception:
                        old_calculation_payload = {}

            calculation_payload = old_calculation_payload or legacy_calculation_payload or {}

            write_json_with_backup(
                target_layout_path,
                normalized_config,
                backup_dir=registry.resolve_backup_dir({"id": form_id}),
                keep_count=12,
            )
            write_json_with_backup(
                target_calculation_path,
                calculation_payload,
                backup_dir=external_path(os.path.join("data", "backups", "form_calculations", form_id).replace("\\", "/")),
                keep_count=12,
            )

            form_record["layout_relative_path"] = target_layout_relative_path
            form_record["layout_path_mode"] = "external"
            migrated.append(form_id)

        normalized_registry_payload = registry._normalize_registry_payload(registry_payload)
        registry._write_registry_payload(normalized_registry_payload)
        return {
            "migrated": migrated,
            "skipped": skipped,
            "count": len(migrated),
        }

    def _get_observed_header_roles(self, header_fields):
        observed_roles = set()
        for field in header_fields if isinstance(header_fields, list) else []:
            if not isinstance(field, dict):
                continue
            field_id = str(field.get("id", "")).strip()
            role_name = resolve_header_field_role(field_id, field.get("role"))
            if role_name:
                observed_roles.add(role_name)
        return observed_roles

    def _get_observed_row_roles(self, row_fields, section_name):
        observed_roles = set()
        for field in row_fields if isinstance(row_fields, list) else []:
            if not isinstance(field, dict):
                continue
            field_id = str(field.get("id", "")).strip()
            role_name = resolve_row_field_role(section_name, field_id, field.get("role"))
            if role_name:
                observed_roles.add(role_name)
        return observed_roles

    def validate_editor_payload_preserves_required_structure(self, payload, payload_details):
        if not isinstance(payload, dict):
            raise ValueError("Layout editor payload must be a JSON object.")

        mode = str((payload_details or {}).get("mode") or "full").strip().lower()
        repeating_bindings = self._iter_repeating_section_bindings(payload)
        row_field_keys = {binding["fields_key"] for binding in repeating_bindings}
        mapping_keys = {binding["mapping_key"] for binding in repeating_bindings}
        list_keys = {"sections", "header_fields", *row_field_keys}
        object_keys = {"editor_presets", *mapping_keys}

        for key_name, value in payload.items():
            if key_name in list_keys and not isinstance(value, list):
                raise ValueError(f"Layout JSON must keep '{key_name}' as a list.")
            if key_name in object_keys and not isinstance(value, dict):
                raise ValueError(f"Layout JSON must keep '{key_name}' as an object.")

        if mode not in {"full", "section"}:
            raise ValueError("Unrecognized editor payload mode.")

    def resolve_editor_text(self, text, base_config=None):
        raw_text = str(text or "").strip()
        if not raw_text:
            raise ValueError("Editor is empty.")

        try:
            payload = json.loads(raw_text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"Syntax error at line {exc.lineno}, column {exc.colno}: {exc.msg}") from exc

        normalized_payload, payload_details = self._normalize_editor_payload(payload, base_config=base_config)
        self.validate_editor_payload_preserves_required_structure(normalized_payload, payload_details)
        active_form_info = None
        try:
            active_form_info = self.service.get_active_form_info()
        except Exception:
            pass
        normalized_config = self.normalize_config(normalized_payload, form_info=active_form_info)
        self.validate_config(normalized_config)
        return normalized_config, payload_details

    def load_current_config(self):
        form_info = self.service.get_active_form_info()
        config, source_path = self.service.load_form(form_info=form_info)
        config = self.normalize_config(config, form_info=form_info)
        self.validate_config(config)
        self.current_source_path = source_path
        self.current_save_path = form_info.get("save_path", source_path)
        self.is_dirty = False
        return config, source_path, form_info

    def load_current_text(self):
        form_info = self.service.get_active_form_info()
        text, _source_path = self.service.load_form_text(form_info=form_info)
        return text

    def load_form_config(self, form_id):
        form_info = self.service.get_form_info(form_id)
        config, source_path = self.service.load_form(form_info=form_info)
        config = self.normalize_config(config, form_info=form_info)
        self.validate_config(config)
        self.current_source_path = source_path
        self.current_save_path = form_info.get("save_path", source_path)
        self.is_dirty = False
        return config, source_path, form_info

    def load_form_text(self, form_id):
        form_info = self.service.get_form_info(form_id)
        text, _source_path = self.service.load_form_text(form_info=form_info)
        return text

    def load_default_config(self):
        config, source_path = self.service.load_default()
        active_form_info = self.service.get_active_form_info()
        config = self.normalize_config(config, form_info=active_form_info)
        self.validate_config(config)
        self.current_source_path = source_path
        self.current_save_path = active_form_info.get("save_path", source_path)
        self.is_dirty = False
        return config, source_path, active_form_info

    def load_default_text(self):
        text, _source_path = self.service.load_default_text()
        return text

    def save_config(self, config, form_info=None):
        resolved_form_info = dict(form_info) if isinstance(form_info, dict) else self.service.get_active_form_info()
        config = self.normalize_config(config, form_info=resolved_form_info)
        self.validate_config(config)
        resolved_form_info = dict(form_info) if isinstance(form_info, dict) else self.service.get_active_form_info()
        backup_info = self.service.save_config(config, form_info=resolved_form_info)
        self.current_source_path = resolved_form_info.get("save_path", self.current_source_path)
        self.current_save_path = resolved_form_info.get("save_path", self.current_save_path)
        self.is_dirty = False
        return backup_info

    def save_config_text(self, text, config=None, form_info=None):
        raw_text = str(text or "")
        resolved_form_info = dict(form_info) if isinstance(form_info, dict) else self.service.get_active_form_info()
        resolved_config = self.normalize_config(config, form_info=resolved_form_info) if isinstance(config, dict) else self.parse_editor_text(raw_text)
        self.validate_config(resolved_config)
        backup_info = self.service.save_config_text(raw_text, form_info=resolved_form_info)
        self.current_source_path = resolved_form_info.get("save_path", self.current_source_path)
        self.current_save_path = resolved_form_info.get("save_path", self.current_save_path)
        self.is_dirty = False
        return backup_info

    def activate_form(self, form_id):
        form_info = self.service.activate_form(form_id)
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        self.is_dirty = False
        return form_info

    def create_form_from_config(self, name, config, description="", activate=False):
        form_id = self.service.registry.canonical_form_id(name) or self.service.registry.normalize_form_id(name)
        form_info = {"id": form_id, "name": name, "description": description}
        config = self.normalize_config(config, form_info=form_info)
        config = self.service.registry._ensure_calculation_metadata(config, form_id)
        self.validate_config(config)
        form_info = self.service.create_form(name, config, description=description, activate=activate)
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        self.is_dirty = False
        return form_info

    def _build_blank_row_fields(self, section_name, required_roles):
        row_fields = []
        for role_name in required_roles:
            field_id = get_default_row_field_id(section_name, role_name) or f"{section_name}_{role_name}"
            field_config = {
                "id": field_id,
                "label": role_name.replace("_", " ").title(),
                "widget": "entry",
                "width": 12,
                "role": role_name,
                "user_input": True,
            }
            if section_name == "production":
                field_config["open_row_trigger"] = True
            row_fields.append(field_config)
        return row_fields

    def _build_blank_mapping(self, mapping_name, row_fields):
        columns = {}
        for index, field in enumerate(row_fields, start=1):
            field_id = str(field.get("id", "")).strip()
            if not field_id:
                continue
            columns[field_id] = self._normalize_mapping_column_config(
                mapping_name,
                field_id,
                {"column": f"COL_{index}"},
                row_fields=row_fields,
            )
        return {
            "start_row": 1,
            "max_rows": DEFAULT_MAPPING_MAX_ROWS,
            "columns": columns,
        }

    def build_blank_form_config(self):
        default_config = self._get_default_config_template()
        active_form_info = None
        try:
            active_form_info = self.service.get_active_form_info()
        except Exception:
            pass
        fallback_prefix = active_form_info.get("name", "Disamatic Production Sheet") if active_form_info else "Disamatic Production Sheet"
        blank_config = {
            "template_path": str(default_config.get("template_path", "")),
            "export_prefix": str(default_config.get("export_prefix", fallback_prefix)),
            "header_fields": [],
            "sections": [],
            "editor_presets": {},
            "calculations": {},
        }

        for binding in self._iter_repeating_section_bindings(default_config):
            section_name = binding["fields_key"]
            mapping_name = binding["mapping_key"]
            blank_config[section_name] = []
            blank_config[mapping_name] = {
                "start_row": 1,
                "max_rows": DEFAULT_MAPPING_MAX_ROWS,
                "columns": {},
            }

        return blank_config

    def create_blank_form(self, name, description="", include_header=True, include_production=True, include_downtime=True, activate=False):
        # 1. Start with the minimal config template structure
        config = {
            "template_path": "",
            "export_prefix": str(name).strip(),
            "header_fields": [],
            "sections": [],
            "editor_presets": {},
            "calculations": {},
        }
        
        default_config = self._get_default_config_template()
        
        # 2. Add sections according to selections
        if include_header:
            config["header_fields"] = deepcopy(default_config.get("header_fields", []))
            for s in default_config.get("sections", []):
                if s.get("behavior_profile") == "header":
                    config["sections"].append(deepcopy(s))
                    break
        
        if include_production:
            config["production_row_fields"] = deepcopy(default_config.get("production_row_fields", []))
            config["production_mapping"] = deepcopy(default_config.get("production_mapping", {
                "start_row": 1,
                "max_rows": DEFAULT_MAPPING_MAX_ROWS,
                "columns": {}
            }))
            for s in default_config.get("sections", []):
                if s.get("behavior_profile") == "production":
                    config["sections"].append(deepcopy(s))
                    break
                    
        if include_downtime:
            config["downtime_row_fields"] = deepcopy(default_config.get("downtime_row_fields", []))
            config["downtime_mapping"] = deepcopy(default_config.get("downtime_mapping", {
                "start_row": 1,
                "max_rows": DEFAULT_MAPPING_MAX_ROWS,
                "columns": {}
            }))
            for s in default_config.get("sections", []):
                if s.get("behavior_profile") == "downtime":
                    config["sections"].append(deepcopy(s))
                    break
        
        config["template_path"] = str(default_config.get("template_path", ""))
        if "export_prefix" in default_config:
            config["export_prefix"] = str(default_config["export_prefix"] if default_config["export_prefix"] is not None else "").strip()
        else:
            config["export_prefix"] = str(name).strip()
        
        # 3. Ensure calculation metadata based on name/form_id
        form_id = self.service.registry.canonical_form_id(name) or self.service.registry.normalize_form_id(name)
        config = self.service.registry._ensure_calculation_metadata(config, form_id)
        
        # 4. Validate and save
        self.validate_config(config)
        form_info = self.service.create_form(name, config, description=description, activate=activate)
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        self.is_dirty = False
        return form_info, deepcopy(config)

    def rename_form(self, form_id, name, description=None):
        form_info = self.service.rename_form(form_id, name, description=description)
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        return form_info

    def duplicate_form(self, source_form_id, name, description=None, activate=False):
        form_info = self.service.duplicate_form(source_form_id, name, description=description, activate=activate)
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        self.is_dirty = False
        return form_info

    def delete_form(self, form_id):
        result = self.service.delete_form(form_id)
        self.current_source_path = self.service.config_path
        self.current_save_path = self.service.save_path
        self.is_dirty = False
        return result

    def list_form_dependencies(self, form_id):
        normalized_form_id = str(form_id or "").strip().lower()
        if not normalized_form_id:
            return []

        pending_dir = external_path(os.path.join("data", "pending"))
        if not os.path.isdir(pending_dir):
            return []

        dependent_drafts = []
        for filename in os.listdir(pending_dir):
            if not str(filename).lower().endswith(".json"):
                continue
            draft_path = os.path.join(pending_dir, filename)
            if not os.path.isfile(draft_path):
                continue
            try:
                with open(draft_path, "r", encoding="utf-8") as handle:
                    payload = json.load(handle)
                meta = payload.get("meta") if isinstance(payload, dict) else {}
                if not isinstance(meta, dict):
                    meta = {}
                draft_form_id = str(meta.get("form_id") or "").strip().lower()
                if draft_form_id != normalized_form_id:
                    continue
                saved_at = str(meta.get("saved_at") or "").strip()
                if not saved_at:
                    saved_at = datetime.fromtimestamp(os.path.getmtime(draft_path)).isoformat(timespec="seconds")
                dependent_drafts.append(
                    {
                        "path": draft_path,
                        "filename": str(filename),
                        "saved_at": saved_at,
                        "form_id": draft_form_id,
                        "form_name": str(meta.get("form_name") or self.get_form_info(normalized_form_id).get("name") or normalized_form_id),
                    }
                )
            except Exception:
                continue

        dependent_drafts.sort(key=lambda item: item.get("saved_at") or "", reverse=True)
        return dependent_drafts

    def build_form_dependency_audit(self, form_id):
        normalized_form_id = str(form_id or "").strip().lower()
        if not normalized_form_id:
            return {
                "form_id": "",
                "dependent_drafts": [],
                "draft_count": 0,
                "summary": "No form selected.",
            }

        dependent_drafts = self.list_form_dependencies(normalized_form_id)
        draft_count = len(dependent_drafts)
        if draft_count == 0:
            summary = "No pending Form Loader drafts depend on this form."
        elif draft_count == 1:
            summary = "1 pending Form Loader draft depends on this form."
        else:
            summary = f"{draft_count} pending Form Loader drafts depend on this form."

        return {
            "form_id": normalized_form_id,
            "dependent_drafts": dependent_drafts,
            "draft_count": draft_count,
            "summary": summary,
        }

    def mark_dirty(self):
        self.is_dirty = True

    def mark_clean(self):
        self.is_dirty = False

    def validate_config(self, config):
        if not isinstance(config, dict):
            raise ValueError("Config must be a JSON object.")

        if "export_prefix" not in config:
            raise ValueError("export_prefix is required in layout config.")
        if not isinstance(config.get("export_prefix"), str):
            raise ValueError("export_prefix must be a string.")

        if "sections" in config and not isinstance(config.get("sections"), list):
            raise ValueError("sections must be a list.")

        seen_section_ids = set()
        seen_supported_profiles = {}
        for index, section in enumerate(config.get("sections", []), start=1):
            if not isinstance(section, dict):
                raise ValueError(f"sections item {index} must be an object.")
            missing_section_keys = [key for key in ("id", "name", "fields_key", "section_type", "behavior_profile") if key not in section]
            if missing_section_keys:
                raise ValueError(f"sections item {index} is missing: {', '.join(missing_section_keys)}")
            section_id = str(section.get("id", "")).strip().lower()
            if not section_id:
                raise ValueError(f"sections item {index} has an empty id.")
            if section_id in seen_section_ids:
                raise ValueError(f"sections contains duplicate id '{section_id}'.")
            seen_section_ids.add(section_id)
            section_type = str(section.get("section_type", "")).strip().lower()
            if section_type not in {"single", "repeating"}:
                raise ValueError(f"sections item {index} uses unsupported section_type '{section.get('section_type')}'.")
            behavior_profile = str(section.get("behavior_profile", "")).strip().lower()
            if not behavior_profile:
                raise ValueError(f"sections item {index} has an empty behavior_profile.")
            if behavior_profile in {"header", "production", "downtime"}:
                existing_section = seen_supported_profiles.get(behavior_profile)
                if existing_section is not None:
                    raise ValueError(
                        f"sections contains duplicate supported behavior_profile '{behavior_profile}' in '{existing_section}' and '{section_id}'."
                    )
                seen_supported_profiles[behavior_profile] = section_id

        if "header_fields" in config:
            self.validate_header_fields(config["header_fields"])

        repeating_bindings = self._iter_repeating_section_bindings(config)
        for binding in repeating_bindings:
            section_name = binding["fields_key"]
            if section_name in config:
                self.validate_row_fields(config.get(section_name), section_name)

        if "editor_presets" in config:
            self.validate_editor_presets(config.get("editor_presets"))

        if "calculations" in config:
            self.validate_calculations_metadata(config.get("calculations"))

        for binding in repeating_bindings:
            section_name = binding["fields_key"]
            mapping_name = binding["mapping_key"]
            if mapping_name not in config:
                continue
            row_fields = config.get(section_name, [])
            required_columns = self.get_required_mapping_field_ids(row_fields, section_name)
            allowed_columns = self.get_mapping_field_ids(row_fields)
            self.validate_mapping(
                config[mapping_name],
                mapping_name,
                required_columns,
                allowed_columns,
            )

    def validate_calculations_metadata(self, calculations):
        if not isinstance(calculations, dict):
            raise ValueError("calculations must be an object.")
        companion_relative_path = str(calculations.get("companion_relative_path") or "").strip()
        if not companion_relative_path:
            raise ValueError("calculations.companion_relative_path is required.")
        section_profiles = calculations.get("section_profiles")
        if not isinstance(section_profiles, list):
            raise ValueError("calculations.section_profiles must be a list.")
        seen_sections = set()
        for index, profile in enumerate(section_profiles, start=1):
            if not isinstance(profile, dict):
                raise ValueError(f"calculations.section_profiles item {index} must be an object.")
            section_id = normalize_role_name(profile.get("section_id"))
            if not section_id:
                raise ValueError(f"calculations.section_profiles item {index} has an empty section_id.")
            if section_id in seen_sections:
                raise ValueError(f"calculations.section_profiles has duplicate section_id '{section_id}'.")
            seen_sections.add(section_id)
            calculation_profile = normalize_role_name(profile.get("calculation_profile"))
            if not calculation_profile:
                raise ValueError(f"calculations.section_profiles item {index} has an empty calculation_profile.")

    def get_mapping_field_ids(self, row_fields):
        field_ids = []
        for field in row_fields if isinstance(row_fields, list) else []:
            if not isinstance(field, dict):
                continue
            field_id = str(field.get("id", "")).strip()
            if field_id:
                field_ids.append(field_id)
        return field_ids

    def validate_header_fields(self, header_fields, field_group_name="header_fields"):
        if not isinstance(header_fields, list):
            raise ValueError(f"{field_group_name} must be a list.")

        allowed_widgets = {"entry", "combobox"}
        allowed_states = {"", "normal", "disabled", "readonly"}
        allowed_options_sources = {"", "downtime_codes"}
        seen_ids = set()
        seen_roles = set()
        for index, field in enumerate(header_fields, start=1):
            if not isinstance(field, dict):
                raise ValueError(f"{field_group_name} item {index} must be an object.")
            field_missing = [key for key in ("id", "label", "row", "col") if key not in field]
            if field_missing:
                raise ValueError(f"{field_group_name} item {index} is missing: {', '.join(field_missing)}")
            field_id = str(field.get("id", "")).strip()
            if not field_id:
                raise ValueError(f"{field_group_name} item {index} has an empty id.")
            if field_id in seen_ids:
                raise ValueError(f"{field_group_name} contains duplicate field id '{field_id}'.")
            seen_ids.add(field_id)

            widget_name = str(field.get("widget", "entry") or "entry").strip().lower() or "entry"
            if widget_name not in allowed_widgets:
                raise ValueError(f"{field_group_name} field '{field_id}' has unsupported widget '{widget_name}'.")

            state_name = str(field.get("state", "") or "").strip().lower()
            if state_name not in allowed_states:
                raise ValueError(f"{field_group_name} field '{field_id}' has unsupported state '{state_name}'.")

            options_source_name = str(field.get("options_source", "") or "").strip().lower()
            from app.downtime_codes import get_available_options_sources
            if options_source_name not in get_available_options_sources():
                raise ValueError(
                    f"{field_group_name} field '{field_id}' has unsupported options_source '{options_source_name}'."
                )

            if widget_name == "combobox":
                values = field.get("values")
                if values is not None and not isinstance(values, list):
                    raise ValueError(f"{field_group_name} field '{field_id}' values must be a list.")

            role_name = resolve_header_field_role(field_id, field.get("role"))
            if role_name:
                if role_name in seen_roles:
                    raise ValueError(f"{field_group_name} contains duplicate role '{role_name}'.")
                seen_roles.add(role_name)

    def validate_editor_presets(self, editor_presets):
        if not isinstance(editor_presets, dict):
            raise ValueError("editor_presets must be an object.")

        config_data = self._get_default_config_template()
        supported_keys = {"header_fields"}
        supported_keys.update(binding["fields_key"] for binding in self._iter_repeating_section_bindings(config_data))
        if not supported_keys:
            supported_keys.update(self.EDITOR_PRESET_SECTION_KEYS)

        unknown_keys = [key for key in editor_presets.keys() if key not in supported_keys]
        if unknown_keys:
            raise ValueError(
                "editor_presets contains unsupported sections: "
                f"{', '.join(sorted(str(key) for key in unknown_keys))}"
            )

        if "header_fields" in editor_presets:
            self.validate_header_fields(editor_presets.get("header_fields"), field_group_name="editor_presets.header_fields")

        row_section_names = [binding["fields_key"] for binding in self._iter_repeating_section_bindings(config_data)]
        if not row_section_names:
            row_section_names = list(self.row_field_sections)
        for section_name in row_section_names:
            if section_name in editor_presets:
                self.validate_row_fields(editor_presets.get(section_name), f"editor_presets.{section_name}")

    def get_required_mapping_field_ids(self, row_fields, section_name):
        normalized_section = normalize_row_section_name(section_name)
        required_roles = REQUIRED_MAPPING_ROLES.get(normalized_section, ())
        role_to_field_id = {}
        for field in row_fields if isinstance(row_fields, list) else []:
            if not isinstance(field, dict):
                continue
            field_id = str(field.get("id", "")).strip()
            if not field_id:
                continue
            role_name = resolve_row_field_role(section_name, field_id, field.get("role"))
            if role_name and role_name not in role_to_field_id:
                role_to_field_id[role_name] = field_id

        required_field_ids = []
        for role_name in required_roles:
            field_id = role_to_field_id.get(role_name)
            if field_id:
                required_field_ids.append(field_id)
        return required_field_ids

    def validate_row_fields(self, row_fields, section_name):
        if not isinstance(row_fields, list):
            raise ValueError(f"{section_name} must be a list.")

        allowed_widgets = {"entry", "display", "checkbutton", "combobox"}
        seen_ids = set()
        seen_roles = set()
        for index, field in enumerate(row_fields, start=1):
            if not isinstance(field, dict):
                raise ValueError(f"{section_name} item {index} must be an object.")
            field_missing = [key for key in ("id", "label", "widget") if key not in field]
            if field_missing:
                raise ValueError(f"{section_name} item {index} is missing: {', '.join(field_missing)}")
            field_id = str(field.get("id", "")).strip()
            if not field_id:
                raise ValueError(f"{section_name} item {index} has an empty id.")
            if field_id in seen_ids:
                raise ValueError(f"{section_name} contains duplicate field id '{field_id}'.")
            seen_ids.add(field_id)
            widget_name = str(field.get("widget", "")).strip().lower()
            if widget_name not in allowed_widgets:
                raise ValueError(
                    f"{section_name} field '{field_id}' uses unsupported widget '{field.get('widget')}'."
                )
            raw_values = field.get("values")
            if raw_values is not None and not isinstance(raw_values, list):
                raise ValueError(f"{section_name} field '{field_id}' values must be a list.")
            role_name = resolve_row_field_role(section_name, field_id, field.get("role"))
            if role_name:
                if role_name in seen_roles:
                    raise ValueError(f"{section_name} contains duplicate role '{role_name}'.")
                seen_roles.add(role_name)

    def validate_mapping(self, mapping, mapping_name, required_columns, allowed_columns=None):
        if not isinstance(mapping, dict):
            raise ValueError(f"{mapping_name} must be an object.")
        if "start_row" not in mapping or "columns" not in mapping:
            raise ValueError(f"{mapping_name} must contain start_row and columns.")
        try:
            start_row = int(mapping.get("start_row", 0))
        except (TypeError, ValueError):
            raise ValueError(f"{mapping_name}.start_row must be an integer.")
        if start_row < 1:
            raise ValueError(f"{mapping_name}.start_row must be 1 or greater.")
        if "max_rows" in mapping:
            try:
                max_rows = int(mapping.get("max_rows", 0))
            except (TypeError, ValueError):
                raise ValueError(f"{mapping_name}.max_rows must be an integer.")
            if max_rows < 1:
                raise ValueError(f"{mapping_name}.max_rows must be 1 or greater.")
        if not isinstance(mapping["columns"], dict):
            raise ValueError(f"{mapping_name}.columns must be an object.")
        for field_id, column_config in mapping["columns"].items():
            if isinstance(column_config, dict):
                column_name = str(column_config.get("column", "")).strip()
                if not column_name:
                    # Allow blank mapping columns so forms can be saved before column assignment is complete.
                    continue
                import_transform = str(column_config.get("import_transform", "value") or "value").strip()
                export_transform = str(column_config.get("export_transform", "value") or "value").strip()
                if import_transform not in VALID_IMPORT_TRANSFORMS:
                    raise ValueError(
                        f"{mapping_name}.columns.{field_id} uses unsupported import_transform '{import_transform}'."
                    )
                if export_transform not in VALID_EXPORT_TRANSFORMS:
                    raise ValueError(
                        f"{mapping_name}.columns.{field_id} uses unsupported export_transform '{export_transform}'."
                    )
                continue
            if not str(column_config or "").strip():
                continue

    def create_unique_field_id(self, config, section_name="header_fields"):
        existing_ids = {field.get("id") for field in config.get(section_name, [])}
        normalized_section_name = str(section_name or "header_fields").strip().lower()
        if normalized_section_name == "header_fields":
            prefix = "new_field"
        else:
            prefix = f"new_{normalize_row_section_name(normalized_section_name) or 'field'}_field"
        index = 1
        while True:
            field_id = f"{prefix}_{index}"
            if field_id not in existing_ids:
                return field_id
            index += 1

    def _mapping_name_for_section(self, section_name, config=None):
        normalized_section_name = str(section_name or "").strip()
        config_data = config if isinstance(config, dict) else {}
        sections = config_data.get("sections") if isinstance(config_data.get("sections"), list) else []
        for section in sections:
            if not isinstance(section, dict):
                continue
            if str(section.get("section_type") or "single").strip().lower() != "repeating":
                continue
            fields_key = str(section.get("fields_key") or "").strip()
            mapping_key = str(section.get("mapping_key") or "").strip()
            if fields_key == normalized_section_name and mapping_key:
                return mapping_key

        if normalized_section_name.endswith("_row_fields"):
            return f"{normalized_section_name[:-11]}_mapping"
        if normalized_section_name.endswith("_fields"):
            return f"{normalized_section_name[:-7]}_mapping"
        return f"{normalize_role_name(normalized_section_name) or 'section'}_mapping"

    def _editor_preset_fields(self, config, section_name):
        editor_presets = config.get("editor_presets") if isinstance(config.get("editor_presets"), dict) else {}
        return editor_presets.get(section_name) if isinstance(editor_presets.get(section_name), list) else []

    def _find_preset_field(self, config, section_name, template_field_id):
        normalized_template_id = str(template_field_id or "").strip()
        if not normalized_template_id:
            return None, ""

        custom_match = next(
            (
                deepcopy(field)
                for field in self._editor_preset_fields(config, section_name)
                if isinstance(field, dict) and str(field.get("id") or "").strip() == normalized_template_id
            ),
            None,
        )
        if custom_match is not None:
            return custom_match, "custom"

        default_config = self._get_default_config_template()
        default_fields = default_config.get(section_name) if isinstance(default_config.get(section_name), list) else []
        default_match = next(
            (
                deepcopy(field)
                for field in default_fields
                if isinstance(field, dict) and str(field.get("id") or "").strip() == normalized_template_id
            ),
            None,
        )
        if default_match is not None:
            return default_match, "default"
        return None, ""

    def _list_available_field_templates(self, config, section_name):
        current_config = self.normalize_config(config) if isinstance(config, dict) else {}
        existing_ids = {
            str(field.get("id") or "").strip()
            for field in current_config.get(section_name, [])
            if isinstance(field, dict)
        }

        templates = []
        seen_ids = set(existing_ids)

        def _append_templates(field_list, source_name):
            for field in field_list if isinstance(field_list, list) else []:
                if not isinstance(field, dict):
                    continue
                field_id = str(field.get("id") or "").strip()
                if not field_id or field_id in seen_ids:
                    continue
                templates.append(
                    {
                        "id": field_id,
                        "label": str(field.get("label") or field_id).strip() or field_id,
                        "source": source_name,
                    }
                )
                seen_ids.add(field_id)

        _append_templates(self._editor_preset_fields(current_config, section_name), "custom")
        default_config = self._get_default_config_template()
        _append_templates(default_config.get(section_name), "default")
        return templates

    def list_available_header_field_templates(self, config):
        return self._list_available_field_templates(config, "header_fields")

    def list_available_row_field_templates(self, config, section_name):
        return self._list_available_field_templates(config, section_name)

    def add_header_field(self, config, insert_index=None):
        field_id = self.create_unique_field_id(config)
        next_row = max((int(field.get("row", 0)) for field in config.get("header_fields", [])), default=-1) + 1
        header_fields = config.setdefault("header_fields", [])
        new_field = {
            "id": field_id,
            "label": field_id.replace("_", " ").title(),
            "row": next_row,
            "col": 0,
            "width": 10,
            "cell": "",
        }
        if insert_index is None:
            header_fields.append(new_field)
        else:
            bounded_index = max(0, min(int(insert_index), len(header_fields)))
            header_fields.insert(bounded_index, new_field)
        return config, f"Added header field '{field_id}'"

    def add_preset_header_field(self, config, template_field_id, insert_index=None):
        normalized_template_id = str(template_field_id or "").strip()
        if not normalized_template_id:
            raise ValueError("Preset field ID is required.")
        template_field, _source_name = self._find_preset_field(config, "header_fields", normalized_template_id)
        if template_field is None:
            raise ValueError(f"Preset field '{normalized_template_id}' was not found for header_fields.")
        existing_ids = {
            str(field.get("id") or "").strip()
            for field in config.get("header_fields", [])
            if isinstance(field, dict)
        }
        if normalized_template_id in existing_ids:
            raise ValueError(f"Field '{normalized_template_id}' already exists in header_fields.")

        existing_roles = {
            resolve_header_field_role(str(field.get("id") or "").strip(), field.get("role"))
            for field in config.get("header_fields", [])
            if isinstance(field, dict)
        }
        existing_roles.discard("")
        template_role = resolve_header_field_role(
            str(template_field.get("id") or "").strip(),
            template_field.get("role"),
        )
        if template_role and template_role in existing_roles:
            template_field.pop("role", None)

        header_fields = config.setdefault("header_fields", [])
        if insert_index is None:
            header_fields.append(template_field)
        else:
            bounded_index = max(0, min(int(insert_index), len(header_fields)))
            header_fields.insert(bounded_index, template_field)
        return config, f"Added preset header field '{normalized_template_id}'"

    def add_row_field(self, config, section_name, insert_index=None):
        section_title = section_name.replace("_", " ").replace(" fields", "").title()
        field_id = self.create_unique_field_id(config, section_name)
        row_fields = config.setdefault(section_name, [])
        new_field = {
            "id": field_id,
            "label": field_id.replace("_", " ").title(),
            "widget": "entry",
            "width": 12,
            "open_row_trigger": True,
            "user_input": True,
        }
        if insert_index is None:
            row_fields.append(new_field)
        else:
            bounded_index = max(0, min(int(insert_index), len(row_fields)))
            row_fields.insert(bounded_index, new_field)
        return config, f"Added {section_title} field '{field_id}'"

    def add_preset_row_field(self, config, section_name, template_field_id, insert_index=None):
        normalized_template_id = str(template_field_id or "").strip()
        if not normalized_template_id:
            raise ValueError("Preset field ID is required.")
        template_field, source_name = self._find_preset_field(config, section_name, normalized_template_id)
        if template_field is None:
            raise ValueError(f"Preset field '{normalized_template_id}' was not found for {section_name}.")
        existing_ids = {
            str(field.get("id") or "").strip()
            for field in config.get(section_name, [])
            if isinstance(field, dict)
        }
        if normalized_template_id in existing_ids:
            raise ValueError(f"Field '{normalized_template_id}' already exists in {section_name}.")
        row_fields = config.setdefault(section_name, [])
        if insert_index is None:
            row_fields.append(template_field)
        else:
            bounded_index = max(0, min(int(insert_index), len(row_fields)))
            row_fields.insert(bounded_index, template_field)
        mapping_name = self._mapping_name_for_section(section_name, config=config)
        default_config = self._get_default_config_template()
        default_mapping = default_config.get(mapping_name) if isinstance(default_config.get(mapping_name), dict) else {}
        default_columns = default_mapping.get("columns") if isinstance(default_mapping.get("columns"), dict) else {}
        mapping = config.setdefault(mapping_name, {})
        columns = mapping.setdefault("columns", {})
        if normalized_template_id in default_columns:
            columns.setdefault(normalized_template_id, deepcopy(default_columns[normalized_template_id]))
        elif source_name == "custom":
            columns.setdefault(
                normalized_template_id,
                self._normalize_mapping_column_config(
                    mapping_name,
                    normalized_template_id,
                    "",
                    row_fields=row_fields,
                ),
            )
        section_title = section_name.replace("_", " ").replace(" fields", "").title()
        return config, f"Added preset {section_title} field '{normalized_template_id}'"

    def replace_header_fields(self, config, header_fields):
        normalized_fields = self._normalize_header_fields(deepcopy(header_fields))
        self.validate_header_fields(normalized_fields)
        config["header_fields"] = normalized_fields
        return config, "Applied header field table edits"

    def replace_row_fields(self, config, section_name, row_fields):
        normalized_fields = []
        source_field_ids = []
        for raw_field in row_fields if isinstance(row_fields, list) else []:
            normalized_field = self._normalize_row_field_entry(raw_field, section_name)
            if not isinstance(normalized_field, dict):
                continue
            normalized_fields.append(normalized_field)
            original_id = str((raw_field or {}).get("_original_id") or normalized_field.get("id") or "").strip()
            source_field_ids.append(original_id)

        self.validate_row_fields(normalized_fields, section_name)
        config[section_name] = normalized_fields

        mapping_name = self._mapping_name_for_section(section_name, config=config)
        mapping = config.setdefault(mapping_name, {})
        existing_columns = mapping.get("columns") if isinstance(mapping.get("columns"), dict) else {}
        rebuilt_columns = {}
        for field, source_field_id in zip(normalized_fields, source_field_ids):
            field_id = str(field.get("id") or "").strip()
            if not field_id:
                continue
            if source_field_id in existing_columns:
                rebuilt_columns[field_id] = deepcopy(existing_columns[source_field_id])
            elif field_id in existing_columns:
                rebuilt_columns[field_id] = deepcopy(existing_columns[field_id])
            else:
                rebuilt_columns[field_id] = self._normalize_mapping_column_config(
                    mapping_name,
                    field_id,
                    "",
                    row_fields=normalized_fields,
                )
        mapping["columns"] = rebuilt_columns
        return config, f"Applied {section_name} table edits"

    def move_header_field(self, config, field_id, direction):
        fields = config.get("header_fields", [])
        current_index = next((index for index, field in enumerate(fields) if field.get("id") == field_id), None)
        if current_index is None:
            raise ValueError(f"Field '{field_id}' was not found.")
        target_index = current_index + direction
        if target_index < 0 or target_index >= len(fields):
            return config, None
        fields[current_index], fields[target_index] = fields[target_index], fields[current_index]
        return config, f"Reordered field '{field_id}'"

    def move_row_field(self, config, section_name, field_id, direction):
        fields = config.get(section_name, [])
        current_index = next((index for index, field in enumerate(fields) if field.get("id") == field_id), None)
        if current_index is None:
            raise ValueError(f"Field '{field_id}' was not found in {section_name}.")
        target_index = current_index + direction
        if target_index < 0 or target_index >= len(fields):
            return config, None
        fields[current_index], fields[target_index] = fields[target_index], fields[current_index]
        return config, f"Reordered field '{field_id}' in {section_name}"

    def remove_header_field(self, config, field_id):
        fields = config.get("header_fields", [])
        updated_fields = [field for field in fields if field.get("id") != field_id]
        if len(updated_fields) == len(fields):
            raise ValueError(f"Field '{field_id}' was not found.")
        config["header_fields"] = updated_fields
        return config, f"Removed field '{field_id}'"

    def remove_row_field(self, config, section_name, field_id):
        fields = config.get(section_name, [])
        updated_fields = [field for field in fields if field.get("id") != field_id]
        if len(updated_fields) == len(fields):
            raise ValueError(f"Field '{field_id}' was not found in {section_name}.")
        config[section_name] = updated_fields
        mapping_name = self._mapping_name_for_section(section_name)
        mapping = config.get(mapping_name)
        columns = mapping.get("columns") if isinstance(mapping, dict) else None
        if isinstance(columns, dict):
            columns.pop(field_id, None)
        return config, f"Removed field '{field_id}' from {section_name}"

    def update_header_field(
        self,
        config,
        field_id,
        updated_field_id,
        label_value,
        row_value,
        col_value,
        cell_value,
        width_value,
        readonly_value,
        default_value,
        role_value,
        import_enabled_value,
        export_enabled_value,
        widget_value,
        state_value,
        options_source_value,
        values_value,
    ):
        if not field_id:
            raise ValueError("Field ID is missing.")
        target_field = None
        for field in config.get("header_fields", []):
            if field.get("id") == field_id:
                target_field = field
                break
        if target_field is None:
            raise ValueError(f"Field '{field_id}' was not found.")

        renamed_field_id = str(updated_field_id or "").strip()
        if not renamed_field_id:
            raise ValueError("Field ID cannot be empty.")
        label_text = str(label_value or "").strip()
        if not label_text:
            raise ValueError("Label cannot be empty.")
        widget_name = str(widget_value or target_field.get("widget", "entry") or "entry").strip().lower() or "entry"
        if widget_name not in {"entry", "combobox"}:
            raise ValueError(f"Unsupported widget type '{widget_name}'.")

        state_name = str(state_value or target_field.get("state", "") or "").strip().lower()
        if state_name not in {"", "normal", "disabled", "readonly"}:
            raise ValueError(f"Unsupported state '{state_name}'.")

        options_source_name = str(options_source_value or target_field.get("options_source", "") or "").strip().lower()
        from app.downtime_codes import get_available_options_sources
        if options_source_name not in get_available_options_sources():
            raise ValueError(f"Unsupported options_source '{options_source_name}'.")

        row = int(str(row_value).strip())
        col = int(str(col_value).strip())
        width = int(str(width_value).strip())
        cell = str(cell_value).strip()
        default_text = str(default_value)
        for field in config.get("header_fields", []):
            if field is target_field:
                continue
            if str(field.get("id") or "").strip() == renamed_field_id:
                raise ValueError(f"Field ID '{renamed_field_id}' is already in use.")
        target_field["id"] = renamed_field_id
        target_field["label"] = label_text
        target_field["row"] = row
        target_field["col"] = col
        if widget_name == "entry":
            target_field.pop("widget", None)
        else:
            target_field["widget"] = widget_name
        target_field["import_enabled"] = self._normalize_bool_value(import_enabled_value, default=True)
        target_field["export_enabled"] = self._normalize_bool_value(export_enabled_value, default=True)
        normalized_role = normalize_role_name(role_value)
        if normalized_role:
            target_field["role"] = normalized_role
        else:
            target_field.pop("role", None)
        if state_name:
            target_field["state"] = state_name
        else:
            target_field.pop("state", None)
        if widget_name == "combobox":
            if options_source_name:
                target_field["options_source"] = options_source_name
            else:
                target_field.pop("options_source", None)
            self._set_optional_list_field(target_field, "values", values_value)
        else:
            target_field.pop("options_source", None)
            target_field.pop("values", None)
        readonly_enabled = self._normalize_bool_value(readonly_value, default=False)
        if target_field.get("id") == "cast_date":
            target_field["readonly"] = True
            target_field.pop("default", None)
        elif readonly_enabled:
            target_field["width"] = width
            target_field["readonly"] = True
        else:
            target_field["width"] = width
            target_field.pop("readonly", None)
        if target_field.get("id") != "cast_date":
            if cell:
                target_field["cell"] = cell
            else:
                target_field.pop("cell", None)
            if default_text.strip():
                target_field["default"] = default_text
            else:
                target_field.pop("default", None)
        normalized_fields = self._normalize_header_fields(deepcopy(config.get("header_fields", [])))
        self.validate_header_fields(normalized_fields)
        config["header_fields"] = normalized_fields
        return config, f"Updated field '{renamed_field_id}'"

    def update_row_field(self, config, section_name, field_id, field_values):
        if not field_id:
            raise ValueError("Field ID is missing.")
        target_field = None
        for field in config.get(section_name, []):
            if field.get("id") == field_id:
                target_field = field
                break
        if target_field is None:
            raise ValueError(f"Field '{field_id}' was not found in {section_name}.")

        renamed_field_id = str(field_values.get("id", target_field.get("id", field_id)) or "").strip()
        if not renamed_field_id:
            raise ValueError("Field ID cannot be empty.")
        for field in config.get(section_name, []):
            if field is target_field:
                continue
            if str(field.get("id") or "").strip() == renamed_field_id:
                raise ValueError(f"Field ID '{renamed_field_id}' is already in use in {section_name}.")

        widget_name = str(field_values.get("widget", target_field.get("widget", "entry"))).strip().lower()
        if widget_name not in {"entry", "display", "checkbutton", "combobox"}:
            raise ValueError(f"Unsupported widget type '{widget_name}'.")

        label_text = str(field_values.get("label", target_field.get("label", field_id))).strip()
        if not label_text:
            raise ValueError("Label cannot be empty.")

        width_text = str(field_values.get("width", target_field.get("width", ""))).strip()
        width_value = int(width_text) if width_text else 0
        if width_value < 0:
            raise ValueError("Width cannot be negative.")

        target_field["id"] = renamed_field_id
        target_field["label"] = label_text
        target_field["widget"] = widget_name
        if width_value > 0:
            target_field["width"] = width_value
        else:
            target_field.pop("width", None)

        explicit_role = normalize_role_name(field_values.get("role"))
        if explicit_role:
            target_field["role"] = explicit_role
        else:
            target_field.pop("role", None)

        user_input_val = self._normalize_bool_value(field_values.get("user_input"), default=False)
        derived_val = self._normalize_bool_value(field_values.get("derived"), default=False)
        if user_input_val:
            derived_val = False

        self._set_bool_field(target_field, "readonly", field_values.get("readonly"), default=False)
        self._set_bool_field(target_field, "derived", derived_val, default=False)
        self._set_bool_field(target_field, "math_trigger", field_values.get("math_trigger"), default=False)
        self._set_bool_field(target_field, "open_row_trigger", field_values.get("open_row_trigger"), default=False)
        self._set_bool_field(target_field, "user_input", user_input_val, default=False)
        self._set_bool_field(target_field, "expand", field_values.get("expand"), default=False)
        self._set_bool_field(target_field, "bold", field_values.get("bold"), default=False)

        self._set_optional_text_field(target_field, "default", field_values.get("default"))
        self._set_optional_text_field(target_field, "sticky", field_values.get("sticky"))
        self._set_optional_text_field(target_field, "state", field_values.get("state"))
        self._set_optional_text_field(target_field, "options_source", field_values.get("options_source"))
        self._set_optional_text_field(target_field, "bootstyle", field_values.get("bootstyle"))
        if widget_name == "combobox":
            self._set_optional_list_field(target_field, "values", field_values.get("values", target_field.get("values")))
        else:
            target_field.pop("values", None)

        if renamed_field_id != field_id:
            mapping_name = self._mapping_name_for_section(section_name, config=config)
            mapping = config.get(mapping_name)
            columns = mapping.get("columns") if isinstance(mapping, dict) else None
            if isinstance(columns, dict) and field_id in columns:
                if renamed_field_id in columns:
                    raise ValueError(
                        f"Cannot rename field '{field_id}' to '{renamed_field_id}' because {mapping_name}.columns already contains that key."
                    )
                columns[renamed_field_id] = columns.pop(field_id)

        return config, f"Updated field '{renamed_field_id}' in {section_name}"

    def _normalize_bool_value(self, value, default=False):
        if isinstance(value, bool):
            return value
        normalized = str(value or "").strip().lower()
        if normalized in {"1", "true", "yes", "on"}:
            return True
        if normalized in {"0", "false", "no", "off"}:
            return False
        return bool(default)

    def _set_bool_field(self, target_field, key_name, value, default=False):
        if self._normalize_bool_value(value, default=default):
            target_field[key_name] = True
        else:
            target_field.pop(key_name, None)

    def _set_optional_text_field(self, target_field, key_name, value):
        text_value = str(value or "").strip()
        if text_value:
            target_field[key_name] = text_value
        else:
            target_field.pop(key_name, None)

    def _normalize_text_list(self, value):
        if isinstance(value, list):
            return [text for text in (str(item).strip() for item in value) if text]

        raw_text = str(value or "").strip()
        if not raw_text:
            return []

        try:
            parsed_value = json.loads(raw_text)
        except (TypeError, ValueError, json.JSONDecodeError):
            parsed_value = None

        if isinstance(parsed_value, list):
            return [text for text in (str(item).strip() for item in parsed_value) if text]

        return [text for text in (item.strip() for item in re.split(r"[,\n]+", raw_text)) if text]

    def _set_optional_list_field(self, target_field, key_name, value):
        normalized_values = self._normalize_text_list(value)
        if normalized_values:
            target_field[key_name] = normalized_values
        else:
            target_field.pop(key_name, None)

    def update_mapping(self, config, mapping_name, start_row_value, max_rows_value, column_values):
        start_row = int(str(start_row_value).strip())
        max_rows = int(str(max_rows_value).strip())
        mapping = config.get(mapping_name)
        if not isinstance(mapping, dict):
            raise ValueError(f"Mapping '{mapping_name}' was not found.")
        mapping["start_row"] = start_row
        mapping["max_rows"] = max_rows
        mapping_columns = mapping.setdefault("columns", {})
        for key, value in column_values.items():
            if isinstance(value, dict):
                cleaned_value = str(value.get("column", "")).strip()
                if not cleaned_value:
                    mapping_columns.pop(key, None)
                    continue
                import_transform = str(value.get("import_transform", "value") or "value").strip() or "value"
                export_transform = str(value.get("export_transform", "value") or "value").strip() or "value"
                if import_transform not in VALID_IMPORT_TRANSFORMS:
                    raise ValueError(f"Column '{key}' uses unsupported import transform '{import_transform}'.")
                if export_transform not in VALID_EXPORT_TRANSFORMS:
                    raise ValueError(f"Column '{key}' uses unsupported export transform '{export_transform}'.")
                mapping_columns[key] = {
                    "column": cleaned_value,
                    "import_enabled": self._normalize_bool_value(value.get("import_enabled"), default=True),
                    "export_enabled": self._normalize_bool_value(value.get("export_enabled"), default=True),
                    "import_transform": import_transform,
                    "export_transform": export_transform,
                }
                continue

            cleaned_value = str(value).strip()
            if not cleaned_value:
                mapping_columns.pop(key, None)
                continue
            mapping_columns[key] = cleaned_value
        return config, f"Updated mapping '{mapping_name}'"

    def update_template_path(self, config, template_path_value):
        config["template_path"] = str(template_path_value or "").strip()
        return config, "Updated export template path"

    def update_export_prefix(self, config, export_prefix_value):
        config["export_prefix"] = str(export_prefix_value or "").strip()
        return config, "Updated export filename prefix"

    def resolve_template_path(self, template_path_value):
        normalized_path = str(template_path_value or "").strip()
        if not normalized_path:
            return ""

        if os.path.isabs(normalized_path) and os.path.exists(normalized_path):
            return normalized_path

        external_candidate = external_path(normalized_path)
        if os.path.exists(external_candidate):
            return external_candidate

        local_or_resource_candidate = local_or_resource_path(normalized_path)
        if os.path.exists(local_or_resource_candidate):
            return local_or_resource_candidate

        if os.path.exists(normalized_path):
            return os.path.abspath(normalized_path)

        return ""

    def build_form_template_relative_path(self, form_info=None, target_filename=""):
        resolved_form_info = dict(form_info) if isinstance(form_info, dict) else self.service.get_active_form_info()
        form_id = self.service.registry.canonical_form_id(str(resolved_form_info.get("id") or ""))
        if not form_id:
            form_id = "form"

        source_name = str(target_filename or "").strip() or "export_template.xlsx"
        sanitized_name = re.sub(r"[^A-Za-z0-9._-]+", "_", os.path.basename(source_name)).strip("._")
        if not sanitized_name:
            sanitized_name = "export_template.xlsx"
        if "." not in sanitized_name:
            sanitized_name = f"{sanitized_name}.xlsx"

        return os.path.join("data", "forms", form_id, sanitized_name).replace("\\", "/")

    def copy_template_to_active_form(self, template_path_value, form_info=None, target_filename=""):
        source_path = self.resolve_template_path(template_path_value)
        if not source_path:
            raise ValueError("Template path is missing or does not exist.")

        target_relative_path = self.build_form_template_relative_path(
            form_info=form_info,
            target_filename=target_filename or os.path.basename(source_path),
        )
        target_absolute_path = external_path(target_relative_path)
        os.makedirs(os.path.dirname(target_absolute_path), exist_ok=True)

        copied = True
        try:
            copied = not os.path.samefile(source_path, target_absolute_path)
        except OSError:
            copied = True

        if copied:
            shutil.copy2(source_path, target_absolute_path)

        return {
            "relative_path": target_relative_path,
            "absolute_path": target_absolute_path,
            "copied": copied,
        }

    def _versions_file_path(self):
        return external_path(LAYOUT_VERSION_FILE)

    def _load_version_store(self):
        version_file = self._versions_file_path()
        try:
            with open(version_file, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
        except Exception:
            payload = {}
        snapshots = payload.get("snapshots") if isinstance(payload.get("snapshots"), list) else []
        return {"snapshots": snapshots}

    def _save_version_store(self, payload):
        version_file = self._versions_file_path()
        backup_dir = external_path("data/backups/layout_versions")
        write_json_with_backup(version_file, payload, backup_dir=backup_dir, keep_count=30)

    def create_form_snapshot(self, form_info, config, label=""):
        snapshot_label = str(label or "").strip() or "Snapshot"
        form_id = str((form_info or {}).get("id") or "active")
        form_name = str((form_info or {}).get("name") or form_id)
        timestamp = datetime.now().isoformat(timespec="seconds")
        snapshot_id = f"{form_id}:{timestamp}"

        store = self._load_version_store()
        snapshots = list(store.get("snapshots") or [])
        snapshots.append(
            {
                "snapshot_id": snapshot_id,
                "form_id": form_id,
                "form_name": form_name,
                "label": snapshot_label,
                "created_at": timestamp,
                "config": self.normalize_config(deepcopy(config), form_info=form_info),
            }
        )
        snapshots = snapshots[-50:]
        store["snapshots"] = snapshots
        self._save_version_store(store)
        return snapshot_id

    def list_form_snapshots(self, form_id=None):
        target_form_id = str(form_id or "").strip()
        snapshots = list(self._load_version_store().get("snapshots") or [])
        if target_form_id:
            snapshots = [item for item in snapshots if str(item.get("form_id") or "") == target_form_id]
        snapshots.sort(key=lambda item: str(item.get("created_at") or ""), reverse=True)
        return snapshots

    def get_form_snapshot(self, snapshot_id):
        target_snapshot_id = str(snapshot_id or "").strip()
        for snapshot in self.list_form_snapshots():
            if str(snapshot.get("snapshot_id") or "") == target_snapshot_id:
                return snapshot
        raise ValueError("Snapshot was not found.")

    def restore_form_snapshot(self, snapshot_id):
        snapshot = self.get_form_snapshot(snapshot_id)
        form_info = {"id": snapshot.get("form_id"), "name": snapshot.get("form_name")}
        config = self.normalize_config(deepcopy(snapshot.get("config") or {}), form_info=form_info)
        self.validate_config(config)
        return config, snapshot

    def bulk_rename_row_fields(self, config, section_name, find_text, replace_text):
        fields = config.get(section_name)
        if not isinstance(fields, list):
            raise ValueError(f"Section '{section_name}' was not found.")
        find_value = str(find_text or "").strip()
        if not find_value:
            raise ValueError("Find text cannot be empty.")
        replace_value = str(replace_text or "")
        updated_count = 0
        for field in fields:
            if not isinstance(field, dict):
                continue
            label_text = str(field.get("label") or "")
            if find_value in label_text:
                field["label"] = label_text.replace(find_value, replace_value)
                updated_count += 1
        if updated_count == 0:
            raise ValueError("No row-field labels matched the find text.")
        return config, f"Renamed {updated_count} row-field labels in {section_name}"

    def bulk_delete_row_fields(self, config, section_name, contains_text):
        fields = config.get(section_name)
        if not isinstance(fields, list):
            raise ValueError(f"Section '{section_name}' was not found.")
        needle = str(contains_text or "").strip().lower()
        if not needle:
            raise ValueError("Match text cannot be empty.")

        updated_fields = []
        removed_count = 0
        for field in fields:
            if not isinstance(field, dict):
                updated_fields.append(field)
                continue
            field_id = str(field.get("id") or "")
            label_text = str(field.get("label") or "")
            is_match = needle in field_id.lower() or needle in label_text.lower()
            if is_match:
                removed_count += 1
                continue
            updated_fields.append(field)

        if removed_count == 0:
            raise ValueError("No non-protected row fields matched the delete filter.")
        config[section_name] = updated_fields
        return config, f"Deleted {removed_count} row fields from {section_name}"

    def bulk_convert_row_widgets(self, config, section_name, from_widget, to_widget):
        fields = config.get(section_name)
        if not isinstance(fields, list):
            raise ValueError(f"Section '{section_name}' was not found.")
        source_widget = str(from_widget or "").strip().lower()
        target_widget = str(to_widget or "").strip().lower()
        if source_widget not in {"entry", "display", "checkbutton", "combobox"}:
            raise ValueError("Source widget type is not supported.")
        if target_widget not in {"entry", "display", "checkbutton", "combobox"}:
            raise ValueError("Target widget type is not supported.")
        if source_widget == target_widget:
            raise ValueError("Source and target widget types are the same.")

        updated_count = 0
        for field in fields:
            if not isinstance(field, dict):
                continue
            current_widget = str(field.get("widget") or "entry").strip().lower()
            if current_widget == source_widget:
                field["widget"] = target_widget
                if target_widget != "combobox":
                    field.pop("values", None)
                updated_count += 1
        if updated_count == 0:
            raise ValueError(f"No row fields were using widget '{source_widget}'.")
        return config, f"Converted {updated_count} row fields from {source_widget} to {target_widget}"

    def get_field_item_key(self, field_id):
        return f"field:{field_id}"

    def get_row_field_item_key(self, section_name, field_id):
        return f"row_field:{section_name}:{field_id}"

    def get_mapping_item_key(self, mapping_name):
        return f"mapping:{mapping_name}"

    def get_protected_row_field_lookup(self, config):
        return {
            binding["fields_key"]: set()
            for binding in self._iter_repeating_section_bindings(config)
        }

    def build_preview_grid(self, config):
        fields = config.get("header_fields", [])
        max_row = max((int(field.get("row", 0)) for field in fields), default=0)
        max_col = max((int(field.get("col", 0)) for field in fields), default=0)
        field_positions = {}

        for field in fields:
            row = int(field.get("row", 0))
            col = int(field.get("col", 0))
            preview_field = dict(field)
            preview_field["item_key"] = self.get_field_item_key(field.get("id", ""))
            field_positions.setdefault((row, col), []).append(preview_field)

        cells = []
        for row in range(max_row + 1):
            for col in range(max_col + 1):
                fields_here = field_positions.get((row, col), [])
                cells.append(
                    {
                        "row": row,
                        "col": col,
                        "fields": fields_here,
                        "item_keys": [field["item_key"] for field in fields_here if field.get("item_key")],
                    }
                )

        row_sections = []
        repeating_bindings = self._iter_repeating_section_bindings(config)
        for binding in repeating_bindings:
            section_id = str(binding.get("section_id") or "").strip() or str(binding.get("fields_key") or "")
            section_name = str(binding.get("fields_key") or "")
            protected_ids = self.get_protected_row_field_lookup(config).get(section_name, set())
            preview_fields = []
            for field in config.get(section_name, []):
                preview_field = dict(field)
                preview_field["item_key"] = self.get_row_field_item_key(section_name, field.get("id", ""))
                preview_field["protected"] = str(field.get("id", "")).strip() in protected_ids
                preview_fields.append(preview_field)
            row_sections.append(
                {
                    "section_id": section_id,
                    "section_name": section_name,
                    "title": self.get_section_name(section_id, config=config, fallback_name=section_name),
                    "description": self.get_section_info(section_id, config=config).get("description", ""),
                    "fields": preview_fields,
                }
            )

        return {
            "field_count": len(fields),
            "max_row": max_row,
            "max_col": max_col,
            "cells": cells,
            "row_sections": row_sections,
        }

    def _build_import_export_cache_key(self, config):
        config_data = self.normalize_config(config)
        template_path = str(config_data.get("template_path") or "").strip()
        repeating_sections = self._iter_repeating_section_bindings(config_data)
        repeating_payload = {}
        for binding in repeating_sections:
            section_name = binding["fields_key"]
            mapping_name = binding["mapping_key"]
            repeating_payload[section_name] = [
                str((field or {}).get("id") or "")
                for field in list(config_data.get(section_name) or [])
                if isinstance(field, dict)
            ]
            repeating_payload[mapping_name] = config_data.get(mapping_name) if isinstance(config_data.get(mapping_name), dict) else {}
        payload = {
            "template_path": template_path,
            "template_signature": self._build_template_file_signature(template_path),
            "repeating_sections": repeating_payload,
        }
        return json.dumps(payload, sort_keys=True, default=str)

    def _build_template_file_signature(self, template_path):
        normalized_path = str(template_path or "").strip()
        if not normalized_path:
            return ""
        resolved_path = self.resolve_template_path(normalized_path)
        if not resolved_path:
            return f"{normalized_path}|missing"
        try:
            stat_info = os.stat(resolved_path)
        except OSError:
            return f"{normalized_path}|missing"
        return f"{normalized_path}|{resolved_path}|{int(stat_info.st_mtime)}|{int(stat_info.st_size)}"

    def _build_template_workbook_stats(self, template_path):
        normalized_path = str(template_path or "").strip()
        if not normalized_path:
            return {
                "exists": False,
                "mode": "none",
                "sheet_count": 0,
                "sheet_names": [],
                "sampled_rows": 0,
                "non_empty_rows": 0,
            }

        cache_key = self._build_template_file_signature(normalized_path)
        resolved_path = self.resolve_template_path(normalized_path)
        cached_stats = self._template_workbook_stats_cache.get(cache_key)
        if isinstance(cached_stats, dict):
            return deepcopy(cached_stats)

        if "|missing" in cache_key:
            stats = {
                "exists": False,
                "mode": "missing",
                "sheet_count": 0,
                "sheet_names": [],
                "sampled_rows": 0,
                "non_empty_rows": 0,
            }
            self._template_workbook_stats_cache[cache_key] = deepcopy(stats)
            return stats

        if load_workbook is None:
            stats = {
                "exists": True,
                "mode": "openpyxl-unavailable",
                "sheet_count": 0,
                "sheet_names": [],
                "sampled_rows": 0,
                "non_empty_rows": 0,
            }
            self._template_workbook_stats_cache[cache_key] = deepcopy(stats)
            return stats

        sheet_names = []
        sampled_rows = 0
        non_empty_rows = 0
        workbook = None
        try:
            workbook = load_workbook(resolved_path, read_only=True, data_only=True)
            worksheets = list(workbook.worksheets)
            for worksheet in worksheets:
                sheet_names.append(str(worksheet.title))
                for row_values in worksheet.iter_rows(min_row=1, max_row=150, values_only=True):
                    sampled_rows += 1
                    if any(cell_value not in (None, "") for cell_value in row_values):
                        non_empty_rows += 1
        except Exception:
            stats = {
                "exists": True,
                "mode": "stream-read-only-failed",
                "sheet_count": 0,
                "sheet_names": [],
                "sampled_rows": 0,
                "non_empty_rows": 0,
            }
            self._template_workbook_stats_cache[cache_key] = deepcopy(stats)
            return stats
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass

        stats = {
            "exists": True,
            "mode": "stream-read-only",
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "sampled_rows": sampled_rows,
            "non_empty_rows": non_empty_rows,
        }
        self._template_workbook_stats_cache[cache_key] = deepcopy(stats)
        if len(self._template_workbook_stats_cache) > 8:
            self._template_workbook_stats_cache = {
                latest_key: self._template_workbook_stats_cache[latest_key]
                for latest_key in list(self._template_workbook_stats_cache.keys())[-8:]
            }
        return stats

    def build_import_export_metadata(self, config):
        config_data = self.normalize_config(config)
        cache_key = self._build_import_export_cache_key(config_data)
        cached_payload = self._import_export_metadata_cache.get(cache_key)
        if isinstance(cached_payload, dict):
            return deepcopy(cached_payload)

        def _mapping_stats(mapping_name, row_section_name):
            mapping = config_data.get(mapping_name) if isinstance(config_data.get(mapping_name), dict) else {}
            columns = mapping.get("columns") if isinstance(mapping.get("columns"), dict) else {}
            row_fields = config_data.get(row_section_name) if isinstance(config_data.get(row_section_name), list) else []

            mapped_columns = 0
            import_enabled_columns = 0
            export_enabled_columns = 0
            transform_overrides = 0

            for raw_value in columns.values():
                if isinstance(raw_value, dict):
                    column_value = str(raw_value.get("column") or "").strip()
                    import_enabled = self._normalize_bool_value(raw_value.get("import_enabled"), default=True)
                    export_enabled = self._normalize_bool_value(raw_value.get("export_enabled"), default=True)
                    import_transform = str(raw_value.get("import_transform") or "value").strip().lower()
                    export_transform = str(raw_value.get("export_transform") or "value").strip().lower()
                else:
                    column_value = str(raw_value or "").strip()
                    import_enabled = True
                    export_enabled = True
                    import_transform = "value"
                    export_transform = "value"

                if column_value:
                    mapped_columns += 1
                if import_enabled:
                    import_enabled_columns += 1
                if export_enabled:
                    export_enabled_columns += 1
                if import_transform != "value" or export_transform != "value":
                    transform_overrides += 1

            return {
                "field_count": len(row_fields),
                "start_row": int(mapping.get("start_row", 1) or 1),
                "max_rows": int(mapping.get("max_rows", DEFAULT_MAPPING_MAX_ROWS) or DEFAULT_MAPPING_MAX_ROWS),
                "mapped_columns": mapped_columns,
                "import_enabled_columns": import_enabled_columns,
                "export_enabled_columns": export_enabled_columns,
                "transform_overrides": transform_overrides,
            }

        section_stats = {}
        for binding in self._iter_repeating_section_bindings(config_data):
            section_id = str(binding.get("section_id") or binding.get("fields_key") or "").strip() or binding["fields_key"]
            section_stats[section_id] = _mapping_stats(binding["mapping_key"], binding["fields_key"])

        metadata = {
            "template_path": str(config_data.get("template_path") or ""),
            "sections": section_stats,
            "workbook": self._build_template_workbook_stats(config_data.get("template_path")),
        }

        production_binding = next(
            (binding for binding in self._iter_repeating_section_bindings(config_data) if str(binding.get("behavior_profile") or "").strip().lower() == "production"),
            None,
        )
        downtime_binding = next(
            (binding for binding in self._iter_repeating_section_bindings(config_data) if str(binding.get("behavior_profile") or "").strip().lower() == "downtime"),
            None,
        )
        metadata["production"] = (
            _mapping_stats(production_binding["mapping_key"], production_binding["fields_key"])
            if production_binding
            else _mapping_stats("production_mapping", "production_row_fields")
        )
        metadata["downtime"] = (
            _mapping_stats(downtime_binding["mapping_key"], downtime_binding["fields_key"])
            if downtime_binding
            else _mapping_stats("downtime_mapping", "downtime_row_fields")
        )

        self._import_export_metadata_cache[cache_key] = deepcopy(metadata)
        if len(self._import_export_metadata_cache) > 12:
            self._import_export_metadata_cache = {
                latest_key: self._import_export_metadata_cache[latest_key]
                for latest_key in list(self._import_export_metadata_cache.keys())[-12:]
            }

        return metadata

    def _normalize_editor_payload(self, payload, base_config=None):
        if not isinstance(payload, dict):
            raise ValueError("JSON editor content must be a full layout JSON object.")

        repeating_bindings = self._iter_repeating_section_bindings(payload)
        dynamic_top_level_keys = {
            binding["fields_key"]
            for binding in repeating_bindings
        }
        dynamic_top_level_keys.update(
            binding["mapping_key"]
            for binding in repeating_bindings
        )
        allowed_keys = set(self.EDITOR_TOP_LEVEL_KEYS) | dynamic_top_level_keys
        unknown_keys = [
            key
            for key in payload.keys()
            if key not in allowed_keys and not str(key).startswith("_")
        ]
        if unknown_keys:
            raise ValueError(
                f"Unknown top-level keys in layout JSON: {', '.join(sorted(str(key) for key in unknown_keys))}"
            )

        dynamic_required_keys = {
            binding["fields_key"]
            for binding in repeating_bindings
        }
        dynamic_required_keys.update(
            binding["mapping_key"]
            for binding in repeating_bindings
        )
        required_keys = tuple(dict.fromkeys((*self.EDITOR_REQUIRED_TOP_LEVEL_KEYS, *sorted(dynamic_required_keys))))
        missing_keys = [key for key in required_keys if key not in payload]
        if missing_keys:
            active_form_info = None
            try:
                active_form_info = self.service.get_active_form_info()
            except Exception:
                pass
            fallback_payload = self.normalize_config(deepcopy(fallback_config), form_info=active_form_info)
            merged_payload = deepcopy(fallback_payload)
            merged_payload.update(payload)

            merged_bindings = self._iter_repeating_section_bindings(merged_payload)
            merged_dynamic_required_keys = {binding["fields_key"] for binding in merged_bindings}
            merged_dynamic_required_keys.update(binding["mapping_key"] for binding in merged_bindings)
            merged_required_keys = tuple(
                dict.fromkeys((*self.EDITOR_REQUIRED_TOP_LEVEL_KEYS, *sorted(merged_dynamic_required_keys)))
            )
            still_missing = [key for key in merged_required_keys if key not in merged_payload]
            if still_missing:
                raise ValueError(
                    "JSON editor expects the full layout object with top-level keys: "
                    f"{', '.join(merged_required_keys)}"
                    f". Optional: {', '.join(self.EDITOR_OPTIONAL_TOP_LEVEL_KEYS)}"
                    f". Missing: {', '.join(still_missing)}"
                )

            warning_message = (
                "Saved using auto-repaired JSON. Missing top-level keys were restored from the active form/default: "
                f"{', '.join(missing_keys)}"
            )
            applied_sections = [key for key in self.EDITOR_TOP_LEVEL_KEYS if key in payload]
            return merged_payload, {
                "mode": "full",
                "applied_sections": applied_sections,
                "auto_filled_missing_keys": list(missing_keys),
                "warning_message": warning_message,
            }

        applied_sections = [key for key in self.EDITOR_TOP_LEVEL_KEYS if key in payload]
        return payload, {"mode": "full", "applied_sections": applied_sections}

    def _infer_list_section(self, payload):
        if not payload:
            return None
        if not all(isinstance(item, dict) for item in payload):
            return None

        if any("row" in item or "col" in item or "cell" in item for item in payload):
            return "header_fields"

        if any("widget" in item or "open_row_trigger" in item or "options_source" in item for item in payload):
            return self._infer_row_field_section(payload)
        return None

    def _infer_row_field_section(self, payload):
        config_data = self._get_default_config_template()
        bindings = self._iter_repeating_section_bindings(config_data)
        if not bindings:
            bindings = [
                {"fields_key": "production_row_fields", "behavior_profile": "production"},
                {"fields_key": "downtime_row_fields", "behavior_profile": "downtime"},
            ]

        section_scores = {}
        for binding in bindings:
            section_name = binding["fields_key"]
            score = 0
            normalized_section = normalize_row_section_name(section_name)
            section_fields = config_data.get(section_name) if isinstance(config_data.get(section_name), list) else []
            protected_ids = {
                str(field.get("id") or "").strip()
                for field in section_fields
                if isinstance(field, dict) and str(field.get("id") or "").strip()
            }
            if not protected_ids:
                protected_ids = self.protected_row_field_ids.get(section_name, set())
            protected_roles = set(PROTECTED_ROW_ROLES.get(normalized_section, set()))
            required_roles = set(REQUIRED_MAPPING_ROLES.get(normalized_section, ()))
            for field in payload:
                field_id = str(field.get("id", "")).strip()
                role_name = resolve_row_field_role(section_name, field_id, field.get("role"))
                if field_id in protected_ids:
                    score += 3
                if role_name in protected_roles:
                    score += 2
                if role_name in required_roles:
                    score += 1
            section_scores[section_name] = score

        highest_score = max(section_scores.values(), default=0)
        if highest_score <= 0:
            return None
        matching_sections = [section_name for section_name, score in section_scores.items() if score == highest_score]
        if len(matching_sections) != 1:
            return None
        return matching_sections[0]

    def _infer_mapping_key(self, payload):
        if not isinstance(payload, dict):
            return None
        if "start_row" not in payload or "columns" not in payload or not isinstance(payload.get("columns"), dict):
            return None

        column_names = {str(column_name).strip() for column_name in payload.get("columns", {}).keys()}
        config_data = self._get_default_config_template()
        mapping_scores = {}
        for binding in self._iter_repeating_section_bindings(config_data):
            section_name = binding["fields_key"]
            mapping_name = binding["mapping_key"]
            section_fields = config_data.get(section_name) if isinstance(config_data.get(section_name), list) else []
            section_field_ids = {
                str(field.get("id") or "").strip()
                for field in section_fields
                if isinstance(field, dict) and str(field.get("id") or "").strip()
            }
            if not section_field_ids:
                section_field_ids = set(self.protected_row_field_ids.get(section_name, set()))
            mapping_scores[mapping_name] = len(column_names & section_field_ids)

        if not mapping_scores:
            mapping_scores = {
                "production_mapping": len(column_names & self.protected_row_field_ids.get("production_row_fields", set())),
                "downtime_mapping": len(column_names & self.protected_row_field_ids.get("downtime_row_fields", set())),
            }
        highest_score = max(mapping_scores.values(), default=0)
        if highest_score <= 0:
            return None
        matching_mappings = [mapping_name for mapping_name, score in mapping_scores.items() if score == highest_score]
        if len(matching_mappings) != 1:
            return None
        return matching_mappings[0]

    def _extract_partial_sections(self, raw_text):
        extracted_sections = {}
        for section_name in self.EDITOR_TOP_LEVEL_KEYS:
            extracted_value = self._extract_named_json_value(raw_text, section_name)
            if extracted_value is not None:
                extracted_sections[section_name] = extracted_value
        return extracted_sections

    def _extract_named_json_value(self, raw_text, key_name):
        pattern = re.compile(rf'"{re.escape(key_name)}"\s*:')
        decoder = json.JSONDecoder()
        for match in pattern.finditer(raw_text):
            value_index = match.end()
            while value_index < len(raw_text) and raw_text[value_index].isspace():
                value_index += 1
            try:
                value, _end_index = decoder.raw_decode(raw_text, idx=value_index)
            except json.JSONDecodeError:
                continue
            return value
        return None

    def build_validation_summary(self, config):
        config_data = self.normalize_config(config)
        errors = []
        warnings = []

        try:
            self.validate_config(config_data)
        except Exception as exc:
            errors.append(str(exc))

        header_fields = config_data.get("header_fields") if isinstance(config_data.get("header_fields"), list) else []
        repeating_bindings = self._iter_repeating_section_bindings(config_data)
        section_field_counts = {"header_fields": len(header_fields)}
        total_fields = len(header_fields)

        def _duplicate_ids(field_list):
            seen_ids = set()
            duplicate_ids = []
            for field in field_list:
                if not isinstance(field, dict):
                    continue
                field_id = str(field.get("id") or "").strip()
                if not field_id:
                    continue
                if field_id in seen_ids and field_id not in duplicate_ids:
                    duplicate_ids.append(field_id)
                seen_ids.add(field_id)
            return duplicate_ids

        field_sections = [("header_fields", header_fields)]
        for binding in repeating_bindings:
            section_name = binding["fields_key"]
            field_list = config_data.get(section_name) if isinstance(config_data.get(section_name), list) else []
            field_sections.append((section_name, field_list))
            section_field_counts[section_name] = len(field_list)
            total_fields += len(field_list)

        for section_name, field_list in field_sections:
            duplicates = _duplicate_ids(field_list)
            if duplicates:
                errors.append(f"{section_name} duplicate ids: {', '.join(duplicates)}")

        for binding in repeating_bindings:
            section_name = binding["fields_key"]
            field_list = config_data.get(section_name) if isinstance(config_data.get(section_name), list) else []
            try:
                self.get_required_mapping_field_ids(field_list, section_name)
            except Exception as exc:
                errors.append(str(exc))

        if total_fields == 0:
            warnings.append("Layout has no editable fields.")

        production_count = section_field_counts.get("production_row_fields", 0)
        downtime_count = section_field_counts.get("downtime_row_fields", 0)

        return {
            "ok": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "stats": {
                "header_fields": len(header_fields),
                "production_row_fields": production_count,
                "downtime_row_fields": downtime_count,
                "sections": section_field_counts,
                "total_fields": total_fields,
            },
        }

    def _resolve_field_role(self, section, field_dict):
        if not isinstance(field_dict, dict):
            return ""
        return str(field_dict.get("role") or "").strip().lower()

    def detect_missing_standard_fields(self, config):
        """
        Detects missing standard roles/fields in the given layout configuration.
        """
        # Header roles
        header_fields = config.get("header_fields", [])
        existing_header_roles = {self._resolve_field_role("header", f) for f in header_fields if isinstance(f, dict)}
        existing_header_roles.discard("")
        
        standard_header_roles = [
            "log_date", "shift_number", "shift_hours", "operator_name", "shift_leader", 
            "production_line", "goal_rate", "total_molds", "shift_start_time", 
            "shift_end_time", "target_time", "cast_date"
        ]
        missing_header = [role for role in standard_header_roles if role not in existing_header_roles]
        
        # Repeating sections
        repeating_bindings = self._iter_repeating_section_bindings(config)
        prod_fields_key = None
        down_fields_key = None
        for binding in repeating_bindings:
            profile = str(binding.get("behavior_profile") or "").strip().lower()
            if profile == "production":
                prod_fields_key = binding["fields_key"]
            elif profile == "downtime":
                down_fields_key = binding["fields_key"]
                
        # Production roles
        standard_prod_roles = [
            "job_order", "part_number", "rate_value", "rate_override_toggle", "mold_count", "duration_minutes"
        ]
        missing_prod = []
        if prod_fields_key:
            existing_prod_roles = {
                self._resolve_field_role("production", f) 
                for f in config.get(prod_fields_key, []) 
                if isinstance(f, dict)
            }
            existing_prod_roles.discard("")
            missing_prod = [role for role in standard_prod_roles if role not in existing_prod_roles]
        else:
            missing_prod = list(standard_prod_roles)
            
        # Downtime roles
        standard_down_roles = [
            "start_clock", "stop_clock", "downtime_code", "cause_text", "duration_minutes"
        ]
        missing_down = []
        if down_fields_key:
            existing_down_roles = {
                self._resolve_field_role("downtime", f) 
                for f in config.get(down_fields_key, []) 
                if isinstance(f, dict)
            }
            existing_down_roles.discard("")
            missing_down = [role for role in standard_down_roles if role not in existing_down_roles]
        else:
            missing_down = list(standard_down_roles)
            
        # Check metadata
        active_form_info = None
        try:
            active_form_info = self.service.get_active_form_info()
        except Exception:
            pass
            
        name_missing = not bool(config.get("export_prefix")) and (not active_form_info or not active_form_info.get("name"))
        desc_missing = not active_form_info or not active_form_info.get("description")
        
        return {
            "header": missing_header,
            "production": missing_prod,
            "downtime": missing_down,
            "metadata": {
                "name_missing": name_missing,
                "description_missing": desc_missing
            }
        }

    def get_column_letter(self, col_idx):
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def get_missing_fields_suggestions(self, config, filename=None):
        """
        Scans filename, form config, and the template path to suggest cell coordinates, 
        column mappings, and names.
        """
        suggestions = {
            "header": {},
            "production": {},
            "downtime": {},
            "form_name": "",
            "description": "",
            "export_prefix": ""
        }
        
        # Analyze filename for Form Name suggestions
        if filename:
            base_name = os.path.splitext(os.path.basename(filename))[0]
            clean_name = base_name.replace("_", " ").replace("-", " ").title()
            suggestions["form_name"] = clean_name
            suggestions["export_prefix"] = clean_name
            suggestions["description"] = f"Imported from {os.path.basename(filename)}"
            
        # Scan Excel template if available
        template_path = config.get("template_path", "")
        if template_path:
            excel_suggestions = self._scan_excel_template_for_roles(template_path)
            suggestions["header"].update(excel_suggestions.get("header", {}))
            suggestions["production"].update(excel_suggestions.get("production", {}))
            suggestions["downtime"].update(excel_suggestions.get("downtime", {}))
            
        # Add fallback/default suggestions for roles that weren't found in Excel
        default_header_mapping = {
            "log_date": ("C4", "Default cell C4"),
            "shift_number": ("C5", "Default cell C5"),
            "shift_hours": ("C6", "Default cell C6"),
            "operator_name": ("C7", "Default cell C7"),
            "shift_leader": ("C8", "Default cell C8"),
            "production_line": ("C9", "Default cell C9"),
            "goal_rate": ("C10", "Default cell C10"),
            "total_molds": ("C11", "Default cell C11"),
            "shift_start_time": ("C12", "Default cell C12"),
            "shift_end_time": ("C13", "Default cell C13"),
            "target_time": ("C14", "Default cell C14"),
            "cast_date": ("C15", "Default cell C15")
        }
        for role, (cell, reason) in default_header_mapping.items():
            if role not in suggestions["header"]:
                suggestions["header"][role] = {"cell": cell, "reason": reason}
                
        default_prod_mapping = {
            "job_order": ("A", 8, "Default column A"),
            "part_number": ("B", 8, "Default column B"),
            "rate_value": ("C", 8, "Default column C"),
            "rate_override_toggle": ("D", 8, "Default column D"),
            "mold_count": ("E", 8, "Default column E"),
            "duration_minutes": ("F", 8, "Default column F")
        }
        for role, (col, start_row, reason) in default_prod_mapping.items():
            if role not in suggestions["production"]:
                suggestions["production"][role] = {"column": col, "start_row": start_row, "reason": reason}
                
        default_down_mapping = {
            "start_clock": ("A", 8, "Default column A"),
            "stop_clock": ("B", 8, "Default column B"),
            "downtime_code": ("C", 8, "Default column C"),
            "cause_text": ("D", 8, "Default column D"),
            "duration_minutes": ("E", 8, "Default column E")
        }
        for role, (col, start_row, reason) in default_down_mapping.items():
            if role not in suggestions["downtime"]:
                suggestions["downtime"][role] = {"column": col, "start_row": start_row, "reason": reason}
                
        return suggestions

    def _scan_excel_template_for_roles(self, template_path):
        resolved_path = self.resolve_template_path(template_path)
        if not resolved_path or not os.path.exists(resolved_path) or load_workbook is None:
            return {"header": {}, "production": {}, "downtime": {}}
            
        header_suggestions = {}
        production_suggestions = {}
        downtime_suggestions = {}
        
        # Keyword mappings
        header_keywords = {
            "log_date": ["date", "log date", "run date", "production date"],
            "shift_number": ["shift", "shift no", "shift number", "shift #"],
            "shift_hours": ["hours", "shift hours", "run hours", "total hours"],
            "operator_name": ["operator", "op", "operator name", "run by"],
            "shift_leader": ["leader", "shift leader", "supervisor"],
            "production_line": ["line", "machine", "line number", "press", "machine #"],
            "goal_rate": ["goal mph", "goal rate", "mph goal", "target mph", "target rate"],
            "total_molds": ["total molds", "molds total", "actual molds"],
            "shift_start_time": ["start time", "shift start", "run start"],
            "shift_end_time": ["end time", "shift end", "run end"],
            "target_time": ["target time", "scheduled time", "minutes target"],
            "cast_date": ["cast date", "cast date #"]
        }
        
        prod_keywords = {
            "job_order": ["order", "shop order", "job order", "order number", "order #"],
            "part_number": ["part", "part number", "part #", "pattern", "pattern #"],
            "rate_value": ["rate", "standard rate", "parts/hour", "mph", "pcs/hr"],
            "rate_override_toggle": ["override", "rate override"],
            "mold_count": ["molds", "actual molds", "mold count", "molds count", "mold"],
            "duration_minutes": ["time", "duration", "run time", "minutes"]
        }
        
        down_keywords = {
            "start_clock": ["start", "start clock", "downtime start", "start time"],
            "stop_clock": ["stop", "stop clock", "downtime stop", "end time"],
            "downtime_code": ["code", "downtime code", "dt code", "code #"],
            "cause_text": ["cause", "reason", "comments", "explanation", "description"],
            "duration_minutes": ["time", "downtime minutes", "duration"]
        }
        
        try:
            wb = load_workbook(resolved_path, data_only=True, read_only=True)
            ws = wb.active
            if ws is None and wb.worksheets:
                ws = wb.worksheets[0]
                
            if ws is not None:
                max_scan_row = min(ws.max_row or 50, 50)
                max_scan_col = min(ws.max_column or 15, 15)
                
                cells_grid = []
                for r in range(1, max_scan_row + 1):
                    row_vals = []
                    for c in range(1, max_scan_col + 1):
                        try:
                            val = ws.cell(row=r, column=c).value
                        except Exception:
                            val = None
                        row_vals.append((c, val))
                    cells_grid.append((r, row_vals))
                    
                for r, row_vals in cells_grid:
                    for c, val in row_vals:
                        if not val or not isinstance(val, str):
                            continue
                        clean_val = val.strip().lower().replace(":", "").replace("-", " ")
                        
                        for role, keywords in header_keywords.items():
                            if role in header_suggestions:
                                continue
                            if any(k in clean_val for k in keywords):
                                target_col = self.get_column_letter(c + 1)
                                target_cell = f"{target_col}{r}"
                                header_suggestions[role] = {
                                    "cell": target_cell,
                                    "reason": f"Found label '{val.strip()}' in cell {self.get_column_letter(c)}{r}"
                                }
                                
                prod_row_candidates = []
                down_row_candidates = []
                
                for r, row_vals in cells_grid[:30]:
                    prod_matches = {}
                    down_matches = {}
                    
                    for c, val in row_vals:
                        if not val or not isinstance(val, str):
                            continue
                        clean_val = val.strip().lower().replace(":", "").replace("-", " ")
                        
                        for role, keywords in prod_keywords.items():
                            if any(k in clean_val for k in keywords):
                                prod_matches[role] = c
                        for role, keywords in down_keywords.items():
                            if any(k in clean_val for k in keywords):
                                down_matches[role] = c
                                
                    if len(prod_matches) >= 2:
                        prod_row_candidates.append((r, prod_matches))
                    if len(down_matches) >= 2:
                        down_row_candidates.append((r, down_matches))
                        
                if prod_row_candidates:
                    best_row, best_matches = max(prod_row_candidates, key=lambda x: len(x[1]))
                    for role, col_idx in best_matches.items():
                        col_letter = self.get_column_letter(col_idx)
                        production_suggestions[role] = {
                            "column": col_letter,
                            "start_row": best_row + 1,
                            "reason": f"Found column header in template row {best_row}"
                        }
                        
                if down_row_candidates:
                    best_row, best_matches = max(down_row_candidates, key=lambda x: len(x[1]))
                    for role, col_idx in best_matches.items():
                        col_letter = self.get_column_letter(col_idx)
                        downtime_suggestions[role] = {
                            "column": col_letter,
                            "start_row": best_row + 1,
                            "reason": f"Found column header in template row {best_row}"
                        }
            wb.close()
        except Exception as exc:
            print(f"Excel template scan error: {exc}")
            
        return {
            "header": header_suggestions,
            "production": production_suggestions,
            "downtime": downtime_suggestions
        }

    def inject_fields_into_config(self, config, fields_to_inject):
        """
        Mutates and normalizes layout configuration to inject the selected fields.
        """
        updated_config = deepcopy(config)
        
        if "sections" not in updated_config or not isinstance(updated_config.get("sections"), list):
            updated_config["sections"] = self._normalize_sections(updated_config)
            
        repeating_bindings = self._iter_repeating_section_bindings(updated_config)
        prod_fields_key = "production_row_fields"
        prod_mapping_key = "production_mapping"
        down_fields_key = "downtime_row_fields"
        down_mapping_key = "downtime_mapping"
        
        for binding in repeating_bindings:
            profile = str(binding.get("behavior_profile") or "").strip().lower()
            if profile == "production":
                prod_fields_key = binding["fields_key"]
                prod_mapping_key = binding["mapping_key"]
            elif profile == "downtime":
                down_fields_key = binding["fields_key"]
                down_mapping_key = binding["mapping_key"]
                
        if "header_fields" not in updated_config:
            updated_config["header_fields"] = []
        if prod_fields_key not in updated_config:
            updated_config[prod_fields_key] = []
        if down_fields_key not in updated_config:
            updated_config[down_fields_key] = []
            
        if prod_mapping_key not in updated_config:
            updated_config[prod_mapping_key] = {"start_row": 8, "max_rows": DEFAULT_MAPPING_MAX_ROWS, "columns": {}}
        if down_mapping_key not in updated_config:
            updated_config[down_mapping_key] = {"start_row": 8, "max_rows": DEFAULT_MAPPING_MAX_ROWS, "columns": {}}
            
        for f in fields_to_inject:
            section = f.get("section")
            role = f.get("role")
            field_id = f.get("id")
            label = f.get("label")
            mapping_val = f.get("mapping")
            widget = f.get("widget", "entry")
            
            if section == "header":
                header_fields = updated_config.get("header_fields", [])
                max_row = max((int(x.get("row", 0)) for x in header_fields), default=-1)
                new_row = max_row + 1
                
                if any(x.get("id") == field_id or x.get("role") == role for x in header_fields):
                    continue
                    
                field_entry = {
                    "id": field_id,
                    "label": label,
                    "row": new_row,
                    "col": 0,
                    "width": 10,
                    "role": role,
                    "widget": widget,
                    "import_enabled": True,
                    "export_enabled": True
                }
                if mapping_val:
                    field_entry["cell"] = mapping_val
                if role == "cast_date":
                    field_entry["readonly"] = True
                    
                header_fields.append(field_entry)
                
            elif section == "production":
                prod_fields = updated_config.get(prod_fields_key, [])
                if any(x.get("id") == field_id or x.get("role") == role for x in prod_fields):
                    continue
                    
                field_entry = {
                    "id": field_id,
                    "label": label,
                    "widget": widget,
                    "width": 12,
                    "role": role,
                    "user_input": True
                }
                if role in {"part_number", "rate_value", "mold_count"}:
                    field_entry["math_trigger"] = True
                if role in {"job_order", "part_number", "mold_count"}:
                    field_entry["open_row_trigger"] = True
                if role in {"rate_value", "duration_minutes"}:
                    field_entry["derived"] = True
                    field_entry.pop("user_input", None)
                if role == "rate_value":
                    field_entry["readonly"] = True
                    field_entry["lookup_source"] = "part_number_rate"
                    field_entry["lookup_key_role"] = "part_number"
                    field_entry["override_toggle_role"] = "rate_override_toggle"
                if role == "rate_override_toggle":
                    field_entry["toggle_target_role"] = "rate_value"
                    field_entry["default"] = False
                    field_entry["widget"] = "checkbutton"
                if role == "duration_minutes":
                    field_entry["widget"] = "display"
                    field_entry["default"] = "0 min"
                    field_entry["sticky"] = "e"
                    field_entry["bold"] = True
                    
                prod_fields.append(field_entry)
                
                if mapping_val:
                    mapping_cols = updated_config[prod_mapping_key].setdefault("columns", {})
                    mapping_cols[field_id] = {
                        "column": mapping_val,
                        "import_enabled": True,
                        "export_enabled": True,
                        "import_transform": "value",
                        "export_transform": "value"
                    }
                    
            elif section == "downtime":
                down_fields = updated_config.get(down_fields_key, [])
                if any(x.get("id") == field_id or x.get("role") == role for x in down_fields):
                    continue
                    
                field_entry = {
                    "id": field_id,
                    "label": label,
                    "widget": widget,
                    "width": 12,
                    "role": role,
                    "user_input": True
                }
                if role in {"start_clock", "stop_clock"}:
                    field_entry["math_trigger"] = True
                if role in {"start_clock", "stop_clock", "downtime_code", "cause_text"}:
                    field_entry["open_row_trigger"] = True
                if role == "duration_minutes":
                    field_entry["derived"] = True
                    field_entry.pop("user_input", None)
                    field_entry["widget"] = "display"
                    field_entry["default"] = "0 min"
                    field_entry["sticky"] = "e"
                    field_entry["bold"] = True
                    field_entry["bootstyle"] = "danger"
                if role == "downtime_code":
                    field_entry["widget"] = "combobox"
                    field_entry["state"] = "readonly"
                    field_entry["options_source"] = "downtime_codes"
                    field_entry["width"] = 18
                if role == "cause_text":
                    field_entry["width"] = 24
                    field_entry["expand"] = True
                    field_entry["sticky"] = "ew"
                    
                down_fields.append(field_entry)
                
                if mapping_val:
                    mapping_cols = updated_config[down_mapping_key].setdefault("columns", {})
                    mapping_cols[field_id] = {
                        "column": mapping_val,
                        "import_enabled": True,
                        "export_enabled": True,
                        "import_transform": "value",
                        "export_transform": "value"
                    }
                    
        return self.normalize_config(updated_config)
